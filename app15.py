from flask import Flask, render_template, request, redirect, session, url_for, flash, jsonify, send_file
from functools import wraps
import mysql.connector
import random
import string
import pandas as pd
from io import BytesIO
from collections import defaultdict
from fpdf import FPDF
import openpyxl
import re
from openpyxl import load_workbook
import os
from werkzeug.utils import secure_filename
import pandas as pd
from io import BytesIO
from flask import send_file
import tempfile
from openpyxl import Workbook



# Khởi tạo Flask app
app = Flask(__name__)
app.secret_key = 'secret123'  # Khóa bảo mật session. Hãy dùng một khóa mạnh hơn trong môi trường production!

# Hàm kết nối đến cơ sở dữ liệu MySQL cho các bảng chung
def get_db_connection():
    return mysql.connector.connect(
        host='localhost',
        user='root',
        password='',
        database='test', # Database chứa info_data, study_data, rules_data, bang_tong_ket
        charset='utf8mb4' # Thêm charset để hỗ trợ tiếng Việt tốt hơn
    )

# Cấu hình cơ sở dữ liệu cho bảng phan_cong
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': '',  # để trống nếu không có mật khẩu
    'database': 'test', # Sửa thành 'test' để đồng nhất
    'charset': 'utf8mb4'
}

# Cấu hình cơ sở dữ liệu cho bảng accounts
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'test',
    'charset': 'utf8mb4'
}

def insert_schedule(data):
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor()
    for row in data:
        khoi = row['khoi']
        tuan = row['tuan']
        lop = row['from']
        lop_truc = row['to']
        cursor.execute('''
            INSERT INTO phan_cong (khoi, tuan, lop, lop_truc)
            VALUES (%s, %s, %s, %s)
        ''', (khoi, tuan, lop, lop_truc))
    conn.commit()
    cursor.close()
    conn.close()

def update_schedule(data):
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE phan_cong
        SET lop = %s, lop_truc = %s
        WHERE khoi = %s AND tuan = %s
    ''', (data['from'], data['to'], data['khoi'], data['tuan']))
    conn.commit()
    cursor.close()
    conn.close()

def clear_all_schedule():
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM phan_cong')
    conn.commit()
    cursor.close()
    conn.close()

def get_all_schedule():
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)
    cursor.execute('SELECT * FROM phan_cong ORDER BY khoi, tuan')
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    return results

# Hàm tạo mật khẩu ngẫu nhiên theo định dạng DDD@DDD
def generate_specific_password():
    part1 = ''.join(random.choices(string.digits, k=3))
    part2 = ''.join(random.choices(string.digits, k=3))
    return f"{part1}@{part2}"

# Hàm cập nhật lop_truc từ bảng phan_cong_truc
def update_lop_truc_data():
    conn = None
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        # Kiểm tra dữ liệu trong phan_cong_truc
        cursor.execute("SELECT COUNT(*) FROM phan_cong_truc")
        count = cursor.fetchone()[0]
        if count == 0:
            return "❌ Không có dữ liệu trong bảng phan_cong_truc để cập nhật."

        sql_query = """
        UPDATE accounts AS a
        JOIN phan_cong_truc AS pct ON a.tuan = pct.tuan AND a.lop = pct.lop
        SET a.lop_truc = pct.lop_truc
        """
        cursor.execute(sql_query)
        conn.commit()
        print(f"Updated {cursor.rowcount} records in accounts with lop_truc.")  # Debug
        return f"✅ Đã cập nhật {cursor.rowcount} bản ghi 'Lớp Trực'."
    except mysql.connector.Error as err:
        print(f"Error updating lop_truc: {err}")  # Debug
        return f"❌ Lỗi cập nhật 'Lớp Trực': {err}"
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

# Decorator kiểm tra người dùng đã đăng nhập hay chưa
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Trang gốc, chuyển hướng đến /home
@app.route('/')
def root():
    print("Redirecting to home_public")  # Debug log
    return redirect(url_for('home_public'))  # Make sure this matches the function name exactly
# Trang đăng nhập
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM accounts WHERE username=%s AND password=%s", (username, password))
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if user:
            session['username'] = user['username']
            session['Name'] = user['Name']
            session['role'] = user['role']
            session['lop'] = user.get('lop', 'N/A')
            session['tuan'] = user.get('tuan', 'N/A')
            session['lop_truc'] = user.get('lop_truc', 'N/A')
            
            flash(f"Chào mừng {user['username']}! Bạn đã đăng nhập thành công.", 'success')
            if user['role'] == 'admin':
                return redirect(url_for('home'))
            elif user['role'] == 'user':
                return redirect(url_for('user'))
            elif user['role'] == 'viewer':
                return redirect(url_for('viewer'))
            elif user['role'] == 'giamthi':
        # Giám thị sẽ vào trang user
                return redirect(url_for('home'))
            else:
                flash("Vai trò không hợp lệ.", 'error')
                return redirect(url_for('login'))
        
        flash("Sai tài khoản hoặc mật khẩu.", 'error')
        return render_template('login.html', error="Sai tài khoản hoặc mật khẩu")
    
    return render_template('login.html')

# Trang chính
@app.route('/home', methods=['GET', 'POST'])
@login_required
def home():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST' and 'set_week' in request.form:
        if session.get('role') == 'admin':
            selected_week = request.form.get('week_select')
            session['tuan'] = selected_week
            flash(f"Tuần đã được đặt thành {selected_week}.", 'info')
            return redirect(url_for('home'))
        else:
            flash("Bạn không có quyền thay đổi tuần hiển thị.", 'error')
            return redirect(url_for('home'))

    if request.method == 'POST' and ('delete_data' in request.form or 'update_data' in request.form):
        if session.get('role') != 'admin':
            flash("Bạn không có quyền thực hiện thao tác này.", 'error')
            return redirect(url_for('home'))
        
        data_id = request.form.get('data_id')
        data_type = request.form.get('data_type')

        if 'delete_data' in request.form:
            if data_type == 'study':
                cursor.execute("DELETE FROM study_data WHERE id = %s", (data_id,))
                flash(f"Đã xóa dữ liệu học tập ID {data_id}.", 'success')
            elif data_type == 'rules':
                cursor.execute("DELETE FROM rules_data WHERE id = %s", (data_id,))
                flash(f"Đã xóa dữ liệu nội quy ID {data_id}.", 'success')
            conn.commit()

        elif 'update_data' in request.form:
            if data_type == 'study':
                return redirect(url_for('update_study_data', data_id=data_id))
            elif data_type == 'rules':
                return redirect(url_for('update_rules_data', data_id=data_id))

    study_data = []
    rules_data = []
    if session.get('role') == 'admin':
        cursor.execute("SELECT * FROM study_data")
        study_data = cursor.fetchall()

        cursor.execute("SELECT * FROM rules_data")
        rules_data = cursor.fetchall()

    cursor.close()
    conn.close()

    lop = session.get('lop', 'Không xác định')
    tuan = session.get('tuan', 'Chưa thiết lập')
    lop_truc = session.get('lop_truc', 'Chưa thiết lập')

    # Get available weeks and classes for export filter
    conn_filter = get_db_connection()
    cursor_filter = conn_filter.cursor()
    
    cursor_filter.execute("SELECT DISTINCT tuan FROM bang_tong_ket UNION SELECT DISTINCT tuan FROM study_data UNION SELECT DISTINCT tuan FROM rules_data ORDER BY tuan ASC")
    available_export_weeks = [row[0] for row in cursor_filter.fetchall()]
    
    # Sửa truy vấn để lấy lop_truc thay vì lop
    cursor_filter.execute("SELECT DISTINCT lop_truc FROM bang_tong_ket UNION SELECT DISTINCT lop_truc FROM study_data UNION SELECT DISTINCT lop_truc FROM rules_data ORDER BY lop_truc ASC")
    available_export_classes = [row[0] for row in cursor_filter.fetchall()]
    
    cursor_filter.close()
    conn_filter.close()

    return render_template('home.html', study_data=study_data, rules_data=rules_data, lop=lop, tuan=tuan, lop_truc=lop_truc,
                           available_export_weeks=available_export_weeks, available_export_classes=available_export_classes)
# Đăng xuất, xóa session
@app.route('/logout')
@login_required
def logout():
    session.clear()
    flash("Bạn đã đăng xuất.", 'info')
    return redirect(url_for('login'))

# Trang admin, xem danh sách người dùng (info_data)
@app.route('/admin')
@login_required
def admin():
    if session.get('role') != ['admin', 'giamthi']:
        flash("Bạn không có quyền truy cập trang admin.", 'error')
        return redirect(url_for('home'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM info_data")
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('admin.html', data=data)

# Thêm dữ liệu vào bảng info_data
@app.route('/add', methods=['POST'])
@login_required
def add():
    if session.get('role') != 'admin':
        flash("Bạn không có quyền thêm dữ liệu.", 'error')
        return redirect(url_for('home'))

    name = request.form['name'].strip()
    email = request.form['email'].strip()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO info_data (name, email) VALUES (%s, %s)", (name, email))
    conn.commit()
    cursor.close()
    conn.close()
    flash("Đã thêm dữ liệu thành công.", 'success')
    return redirect(url_for('admin'))

# Chỉnh sửa thông tin trong bảng info_data
@app.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit(id):
    if session.get('role') != 'admin':
        flash("Bạn không có quyền chỉnh sửa dữ liệu.", 'error')
        return redirect(url_for('home'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        name = request.form['name'].strip()
        email = request.form['email'].strip()
        cursor.execute("UPDATE info_data SET name=%s, email=%s WHERE id=%s", (name, email, id))
        conn.commit()
        cursor.close()
        conn.close()
        flash(f"Đã cập nhật ID {id} thành công.", 'success')
        return redirect(url_for('admin'))

    cursor.execute("SELECT * FROM info_data WHERE id=%s", (id,))
    user_info = cursor.fetchone()
    cursor.close()
    conn.close()
    if not user_info:
        flash("Không tìm thấy dữ liệu để chỉnh sửa.", 'error')
        return redirect(url_for('admin'))
    return render_template('edit.html', user=user_info)

# Xóa bản ghi trong bảng info_data
@app.route('/delete/<int:id>')
@login_required
def delete(id):
    if session.get('role') != 'admin':
        flash("Bạn không có quyền xóa dữ liệu.", 'error')
        return redirect(url_for('home'))

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM info_data WHERE id=%s", (id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash(f"Đã xóa dữ liệu ID {id} thành công.", 'success')
    return redirect(url_for('admin'))

# Trang người dùng (chỉ người dùng role 'user' mới xem được)


@app.route('/user', methods=['GET', 'POST'])
@login_required
def user():
    if session.get('role') == 'giamthi':
        return redirect(url_for('home'))
    if session.get('role') not in ['admin', 'user', 'viewer']:
        flash("Bạn không có quyền truy cập trang người dùng.", 'error')
        return redirect(url_for('home'))
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    if request.method == 'POST':
        data_id = request.form.get('data_id')
        data_type = request.form.get('data_type')
        user_lop_truc = session.get('lop_truc')
        user_tuan = session.get('tuan')
        cursor.execute("SELECT trangthai FROM bang_tong_ket WHERE lop_truc = %s AND tuan = %s", (user_lop_truc, user_tuan))
        account_status = cursor.fetchone()
        cursor.fetchall()  # Đọc hết kết quả
        trangthai_tongket = account_status['trangthai'] if account_status else 'Chưa tổng kết'
        if trangthai_tongket == 'Đã tổng kết':
            flash("Không thể chỉnh sửa hoặc xóa dữ liệu vì tuần này đã được tổng kết.", 'warning')
            cursor.close()
            conn.close()
            return redirect(url_for('user'))
        if 'delete_data' in request.form:
            if data_type == 'study':
                cursor.execute("DELETE FROM study_data WHERE id = %s", (data_id,))
                flash(f"Đã xóa dữ liệu học tập ID {data_id}.", 'success')
            elif data_type == 'rules':
                cursor.execute("DELETE FROM rules_data WHERE id = %s", (data_id,))
                flash(f"Đã xóa dữ liệu nội quy ID {data_id}.", 'success')
            conn.commit()
            cursor.close()
            conn.close()
            return redirect(url_for('user'))
        elif 'update_data' in request.form:
            if data_type == 'study':
                return redirect(url_for('update_study_data', data_id=data_id))
            elif data_type == 'rules':
                return redirect(url_for('update_rules_data', data_id=data_id))
    
    user_lop = session.get('lop')
    user_tuan_hien_tai = session.get('tuan')
    user_lop_truc = session.get('lop_truc')
    selected_tuan_tong_ket = request.args.get('tong_ket_tuan', type=str)
    view_all = request.args.get('view_all', type=str) == 'true'
    
    # Kiểm tra trạng thái tổng kết
    trangthai_tongket = 'Chưa tổng kết'
    if user_lop_truc and user_tuan_hien_tai:
        cursor.execute("SELECT trangthai FROM bang_tong_ket WHERE lop_truc = %s AND tuan = %s", (user_lop_truc, user_tuan_hien_tai))
        account_status = cursor.fetchone()
        cursor.fetchall()  # Đọc hết kết quả
        trangthai_tongket = account_status['trangthai'] if account_status else 'Chưa tổng kết'
    
    study_data = []
    rules_data = []
    tong_ket_data = []
    available_weeks_for_lop = []
    
    # Lấy danh sách tuần có dữ liệu tổng kết
    cursor.execute("SELECT DISTINCT tuan FROM bang_tong_ket ORDER BY tuan ASC")
    available_weeks_for_lop = [row['tuan'] for row in cursor.fetchall()]
    if not selected_tuan_tong_ket and available_weeks_for_lop and not view_all:
        selected_tuan_tong_ket = available_weeks_for_lop[0]
    
    # Lấy dữ liệu học tập và nội quy (chỉ cho lop_truc của user)
    if user_lop_truc and user_tuan_hien_tai:
        cursor.execute("SELECT * FROM study_data WHERE lop_truc = %s AND tuan = %s", (user_lop_truc, user_tuan_hien_tai))
        study_data = cursor.fetchall()
        cursor.execute("SELECT * FROM rules_data WHERE lop_truc = %s AND tuan = %s", (user_lop_truc, user_tuan_hien_tai))
        rules_data = cursor.fetchall()
    else:
        flash("Thông tin lớp trực hoặc tuần của tài khoản bạn chưa được thiết lập. Vui lòng liên hệ quản trị viên.", 'info')
    
    # Lấy dữ liệu tổng kết cho tất cả các lớp
    query = "SELECT tuan, khoi, lop_truc, tong_diem_hoc_tap, tong_diem_noi_quy, tong_diem_chung FROM bang_tong_ket WHERE 1=1"
    params = []
    if selected_tuan_tong_ket and not view_all:
        query += " AND tuan = %s"
        params.append(selected_tuan_tong_ket)
    query += " ORDER BY tuan ASC, khoi ASC, tong_diem_chung DESC"
    cursor.execute(query, params)
    tong_ket_data = cursor.fetchall()
    if tong_ket_data:
        grouped_data = defaultdict(list)
        for item in tong_ket_data:
            grouped_data[(item['tuan'], item.get('khoi', 'Unknown'))].append(item)
        ranked_data = []
        for tuan_khoi in sorted(grouped_data.keys()):
            current_week_block_data = sorted(grouped_data[tuan_khoi], key=lambda x: x['tong_diem_chung'], reverse=True)
            current_rank = 1
            same_rank_count = 1
            prev_diem = None
            for i, item in enumerate(current_week_block_data):
                if i == 0:
                    item['xep_hang'] = current_rank
                else:
                    if item['tong_diem_chung'] < prev_diem:
                        current_rank += same_rank_count
                        same_rank_count = 1
                    else:
                        same_rank_count += 1
                    item['xep_hang'] = current_rank
                prev_diem = item['tong_diem_chung']
                ranked_data.append(item)
        tong_ket_data = ranked_data
    else:
        flash("Không có dữ liệu tổng kết để hiển thị.", 'warning')
    
    cursor.fetchall()  # Đọc hết kết quả
    cursor.close()
    conn.close()
    return render_template('user.html', study_data=study_data, rules_data=rules_data, lop=user_lop, tuan=user_tuan_hien_tai, lop_truc=user_lop_truc, trangthai_tongket=trangthai_tongket, tong_ket_data=tong_ket_data, selected_tuan_tong_ket=selected_tuan_tong_ket, available_weeks_for_lop=available_weeks_for_lop, view_all=view_all)


# @app.route('/giamthi', methods=['GET', 'POST'])
# @login_required
# def giamthi():
#     if session.get('role') != 'giamthi':
#         flash("Bạn không có quyền truy cập trang giám thị.", 'error')
#         return redirect(url_for('home'))

#     # Giám thị dùng lại giao diện user nhưng có thêm quyền chỉnh sửa nội quy, học tập và xem tổng kết
#     return render_template('giamthi.html',
#                            lop=session.get('lop'),
#                            tuan=session.get('tuan'),
#                            lop_truc=session.get('lop_truc'))

# Trang viewer (role viewer)
@app.route('/viewer')
@login_required
def viewer():
    if session.get('role') != 'viewer':
        flash("Bạn không có quyền truy cập trang xem.", 'error')
        return redirect(url_for('home'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    user_lop = session.get('lop')
    user_tuan = session.get('tuan')
    user_lop_truc = session.get('lop_truc')

    study_data = []
    rules_data = []

    if user_lop and user_tuan:
        cursor.execute("SELECT * FROM study_data WHERE lop = %s AND tuan = %s", (user_lop, user_tuan))
        study_data = cursor.fetchall()

        cursor.execute("SELECT * FROM rules_data WHERE lop = %s AND tuan = %s", (user_lop, user_tuan))
        rules_data = cursor.fetchall()

    cursor.close()
    conn.close()
    return render_template('viewer.html', 
                           study_data=study_data, 
                           rules_data=rules_data, 
                           lop=user_lop, 
                           tuan=user_tuan, 
                           lop_truc=user_lop_truc)

# Trang Học Tập (study_data)

# @app.route('/hoc_tap', methods=['GET', 'POST'])
# @login_required
# def hoc_tap():
#     if session.get('role') not in ['admin', 'user', 'viewer']:
#         flash("Bạn không có quyền truy cập vào mục Học Tập.", 'error')
#         return redirect(url_for('home'))
#     conn = get_db_connection()
#     cursor = conn.cursor(dictionary=True)
#     data = []
#     user_role = session.get('role')
#     user_lop_truc = session.get('lop_truc')
#     user_tuan = session.get('tuan')
    
#     # Lấy danh sách tuần và lớp trực
#     cursor.execute("SELECT DISTINCT tuan FROM study_data ORDER BY tuan ASC")
#     available_weeks = [row['tuan'] for row in cursor.fetchall()]
#     cursor.execute("SELECT DISTINCT lop_truc FROM study_data ORDER BY lop_truc ASC")
#     available_lops = [row['lop_truc'] for row in cursor.fetchall()]
    
#     selected_tuan = request.args.get('tuan', type=str)
#     selected_lop = request.args.get('lop', type=str)
    
#     # Kiểm tra trạng thái tổng kết
#     trangthai_tongket = 'Chưa tổng kết'
#     if user_lop_truc and user_tuan:
#         cursor.execute("SELECT trangthai FROM bang_tong_ket WHERE lop_truc = %s AND tuan = %s", (user_lop_truc, user_tuan))
#         status_row = cursor.fetchone()
#         if status_row:
#             trangthai_tongket = status_row['trangthai']
    
#     query = "SELECT * FROM study_data WHERE 1=1"
#     query_params = []
#     if user_role == 'admin':
#         if selected_tuan:
#             query += " AND tuan = %s"
#             query_params.append(selected_tuan)
#         if selected_lop:
#             query += " AND lop_truc = %s"
#             query_params.append(selected_lop)
#     elif user_role in ['user', 'viewer']:
#         if user_lop_truc and user_tuan:
#             query += " AND lop_truc = %s AND tuan = %s"
#             query_params.append(user_lop_truc)
#             query_params.append(user_tuan)
#         else:
#             flash("Không có thông tin lớp trực hoặc tuần để hiển thị dữ liệu học tập.", 'info')
#     query += " ORDER BY tuan DESC, lop_truc ASC"
#     cursor.execute(query, tuple(query_params))
#     data = cursor.fetchall()
#     cursor.close()
#     conn.close()
#     return render_template('hoc_tap.html', data=data, available_weeks=available_weeks, available_lops=available_lops, selected_tuan=selected_tuan, selected_lop=selected_lop, trangthai_tongket=trangthai_tongket)


@app.route('/hoc_tap', methods=['GET', 'POST'])
@login_required
def hoc_tap():
    if session.get('role') not in ['admin', 'giamthi', 'user', 'viewer']:
        flash("Bạn không có quyền truy cập trang học tập.", 'error')
        return redirect(url_for('home'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    user_lop_truc = session.get('lop_truc')
    user_tuan = session.get('tuan')

    # Kiểm tra trạng thái tổng kết (áp dụng cho user, admin và giamthi)
    trangthai_tongket = 'Chưa tổng kết'
    if user_lop_truc and user_tuan:
        cursor.execute("""
            SELECT trangthai FROM bang_tong_ket 
            WHERE lop_truc = %s AND tuan = %s
        """, (user_lop_truc, user_tuan))
        result = cursor.fetchone()
        cursor.fetchall()
        if result:
            trangthai_tongket = result['trangthai']

    # Xử lý xóa dữ liệu
    if request.method == 'POST':
        if 'delete_data' in request.form and session.get('role') in ['admin', 'giamthi']:
            data_id = request.form.get('data_id')
            cursor.execute("DELETE FROM study_data WHERE id = %s", (data_id,))
            conn.commit()
            flash(f"Đã xóa dữ liệu học tập ID {data_id}.", 'success')
            cursor.close()
            conn.close()
            return redirect(url_for('hoc_tap'))

        elif 'delete_all' in request.form and session.get('role') in ['admin', 'giamthi']:
            password = request.form.get('password')
            if password == '1233':
                cursor.execute("TRUNCATE TABLE study_data")
                conn.commit()
                flash("Đã xóa toàn bộ dữ liệu học tập.", 'success')
            else:
                flash("Mật khẩu không đúng. Vui lòng nhập lại.", 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('hoc_tap'))

    # Lấy dữ liệu học tập
    query = "SELECT * FROM study_data WHERE 1=1"
    params = []

    if session.get('role') == 'user':
        # User chỉ xem dữ liệu của lớp trực và tuần của mình
        if user_lop_truc and user_tuan:
            query += " AND lop_truc = %s AND tuan = %s"
            params.extend([user_lop_truc, user_tuan])
    # Admin & giamthi xem tất cả, không cần lọc

    query += " ORDER BY tuan ASC, lop_truc ASC"
    cursor.execute(query, params)
    study_data = cursor.fetchall()
    cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'hoc_tap.html',
        study_data=study_data,
        trangthai_tongket=trangthai_tongket
    )


@app.route('/delete_hoc_tap_entry/<int:entry_id>', methods=['POST'])
@login_required
def delete_hoc_tap_entry(entry_id):
    user_role = session.get('role')
    current_class_id = session.get('lop')

    if user_role not in ['admin', 'user', 'giamthi']:
        return jsonify({'status': 'error', 'message': 'Bạn không có quyền thực hiện hành động này.'}), 403

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT lop, tuan FROM study_data WHERE id = %s", (entry_id,))
        entry_info = cursor.fetchone()

        if not entry_info:
            return jsonify({'status': 'error', 'message': 'Dữ liệu không tồn tại.'}), 404
        
        entry_lop = entry_info['lop']
        entry_tuan = entry_info['tuan']

        cursor.execute("SELECT trangthai FROM accounts WHERE username = %s AND tuan = %s", 
                      (session.get('username'), entry_tuan))
        account_status = cursor.fetchone()
        if account_status and account_status['trangthai'] == 'Đã tổng kết':
            return jsonify({'status': 'error', 'message': 'Không thể xóa dữ liệu vì tuần này đã được tổng kết.'}), 403

        if user_role == 'user' and entry_lop != current_class_id:
            return jsonify({'status': 'error', 'message': 'Bạn không có quyền xóa dữ liệu của lớp khác.'}), 403

        cursor.execute("DELETE FROM study_data WHERE id = %s", (entry_id,))
        conn.commit()
        flash("Đã xóa dữ liệu học tập thành công.", 'success')
        return jsonify({'status': 'success', 'message': 'Đã xóa dữ liệu học tập thành công.'}), 200
    except mysql.connector.Error as err:
        flash(f"Lỗi khi xóa dữ liệu học tập: {err}", 'error')
        print(f"Error deleting study data entry: {err}")
        return jsonify({'status': 'error', 'message': f'Lỗi khi xóa dữ liệu học tập: {err}'}), 500
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/add_hoc_tap', methods=['GET', 'POST'])
@login_required
def add_hoc_tap():
    # Chỉ cho phép admin, giamthi, user
    if session.get('role') not in ['admin', 'giamthi', 'user']:
        flash("Bạn không có quyền thêm dữ liệu học tập.", 'error')
        return redirect(url_for('hoc_tap'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    user_lop = session.get('lop', '')
    user_tuan = session.get('tuan', '')
    user_lop_truc = ''

    # Lấy lớp trực được phân công nếu là user
    if user_lop and user_tuan:
        cursor.execute("""
            SELECT lop_truc 
            FROM phan_cong_truc 
            WHERE lop = %s AND tuan = %s
        """, (user_lop, user_tuan))
        result = cursor.fetchone()
        user_lop_truc = result['lop_truc'] if result else 'Chưa gán'

    # Lấy danh sách lớp trực:
    # - Admin & Giám thị: thấy tất cả lớp trực
    # - User: chỉ thấy lớp trực của tuần mình
    if session.get('role') in ['admin', 'giamthi']:
        cursor.execute("SELECT DISTINCT lop_truc FROM phan_cong_truc ORDER BY lop_truc")
    else:
        cursor.execute("""
            SELECT DISTINCT lop_truc 
            FROM phan_cong_truc 
            WHERE tuan = %s 
            ORDER BY lop_truc
        """, (user_tuan,))
    available_lop_truc = [row['lop_truc'] for row in cursor.fetchall()]

    # Nếu là user nhưng chưa được phân công
    if session.get('role') == 'user' and (not user_lop_truc or not user_tuan):
        flash("Bạn cần được gán lớp trực và tuần trước khi thêm dữ liệu học tập.", 'error')
        cursor.close()
        conn.close()
        return redirect(url_for('user'))

    # Kiểm tra trạng thái tổng kết
    cursor.execute("""
        SELECT trangthai 
        FROM bang_tong_ket 
        WHERE lop_truc = %s AND tuan = %s
    """, (user_lop_truc, user_tuan))
    status_row = cursor.fetchone()
    if status_row and status_row['trangthai'] == 'Đã tổng kết':
        flash("Tuần này đã được tổng kết. Bạn không thể thêm dữ liệu học tập.", 'warning')
        cursor.close()
        conn.close()
        return redirect(url_for('hoc_tap'))

    # Xử lý POST form
    if request.method == 'POST':
        if session.get('role') in ['admin', 'giamthi']:
            tuan = request.form['tuan'].strip()
            lop_truc = request.form['lop_truc'].strip()
        else:
            tuan = user_tuan
            lop_truc = user_lop_truc

        gio_a = int(request.form['gio_a'] or 0)
        gio_b = int(request.form['gio_b'] or 0)
        gio_c = int(request.form['gio_c'] or 0)
        gio_d = int(request.form['gio_d'] or 0)
        dat_kieu_mau = request.form['dat_kieu_mau']

        # Tính tổng điểm
        tong_diem = gio_a * 5 + gio_b * -5 + gio_c * -15 + gio_d * -25
        tong_diem += 5 if dat_kieu_mau == "Yes" else -10

        try:
            cursor.execute("""
                INSERT INTO study_data 
                (tuan, lop, lop_truc, gio_a, gio_b, gio_c, gio_d, dat_kieu_mau, tong_diem)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (tuan, user_lop, lop_truc, gio_a, gio_b, gio_c, gio_d, dat_kieu_mau, tong_diem))
            conn.commit()
            flash("Đã thêm dữ liệu học tập thành công.", 'success')
        except mysql.connector.Error as err:
            flash(f"Lỗi khi thêm dữ liệu học tập: {err}", 'error')
        finally:
            cursor.close()
            conn.close()

        return redirect(url_for('hoc_tap'))

    cursor.close()
    conn.close()

    return render_template(
        'add_hoc_tap.html',
        user_lop=user_lop,
        user_tuan=user_tuan,
        user_lop_truc=user_lop_truc,
        role=session.get('role'),
        available_lop_truc=available_lop_truc
    )


# Trang Nội Quy
@app.route('/noi_quy', methods=['GET', 'POST'])
@login_required
def noi_quy():
    if session.get('role') not in ['admin', 'user', 'giamthi']:
        flash("Bạn không có quyền truy cập trang nội quy.", 'error')
        return redirect(url_for('home'))
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    user_lop_truc = session.get('lop_truc')
    user_tuan = session.get('tuan')
    
    # Kiểm tra trạng thái tổng kết
    trangthai_tongket = 'Chưa tổng kết'
    if user_lop_truc and user_tuan:
        cursor.execute("SELECT trangthai FROM bang_tong_ket WHERE lop_truc = %s AND tuan = %s", (user_lop_truc, user_tuan))
        result = cursor.fetchone()
        cursor.fetchall()  # Đọc hết kết quả
        if result:
            trangthai_tongket = result['trangthai']
    
    # Xử lý xóa dữ liệu
    if request.method == 'POST':
        if 'delete_data' in request.form and session.get('role') in ['admin', 'giamthi']:
    
            data_id = request.form.get('data_id')
            cursor.execute("DELETE FROM rules_data WHERE id = %s", (data_id,))
            conn.commit()
            flash(f"Đã xóa dữ liệu nội quy ID {data_id}.", 'success')
            cursor.close()
            conn.close()
            return redirect(url_for('noi_quy'))
        elif 'delete_all' in request.form and session.get('role')  in ['admin', 'giamthi']:
            password = request.form.get('password')
            if password == '1233':
                cursor.execute("TRUNCATE TABLE rules_data")
                conn.commit()
                flash("Đã xóa toàn bộ dữ liệu nội quy.", 'success')
            else:
                flash("Mật khẩu không đúng. Vui lòng nhập lại.", 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('noi_quy'))
    
    # Lấy dữ liệu nội quy
    query = "SELECT * FROM rules_data WHERE 1=1"
    params = []
    if session.get('role') == 'user':
        if user_lop_truc and user_tuan:
            query += " AND lop_truc = %s AND tuan = %s"
            params.extend([user_lop_truc, user_tuan])
    query += " ORDER BY tuan ASC, lop_truc ASC"
    cursor.execute(query, params)
    rules_data = cursor.fetchall()
    cursor.fetchall()  # Đọc hết kết quả
    
    cursor.close()
    conn.close()
    return render_template('noi_quy.html', rules_data=rules_data, trangthai_tongket=trangthai_tongket)

@app.route('/edit_noi_quy/<int:rule_id>', methods=['GET', 'POST'])
@login_required
def edit_noi_quy(rule_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Lấy dữ liệu theo ID
    cursor.execute("SELECT * FROM rules_data WHERE id = %s", (rule_id,))
    rule = cursor.fetchone()

    if not rule:
        flash("Không tìm thấy dữ liệu nội quy.", "error")
        return redirect(url_for('noi_quy'))

    if request.method == 'POST':
        # Lấy dữ liệu từ form
        tuan = request.form['tuan']
        lop = request.form['lop']
        vi_pham = request.form['vi_pham']
        diem_tru = int(request.form['diem_tru'] or 0)
        so_luot = int(request.form['so_luot'] or 0)
        hoc_sinh = request.form['hoc_sinh']
        tong_diem = diem_tru * so_luot

        try:
            cursor.execute("""
                UPDATE rules_data
                SET tuan=%s, lop=%s, noi_dung_vi_pham=%s, diem_tru=%s, so_luot_vi_pham=%s,
                    ten_hoc_sinh_vi_pham=%s, tong_diem_vi_pham=%s
                WHERE id=%s
            """, (tuan, lop, vi_pham, diem_tru, so_luot, hoc_sinh, tong_diem, rule_id))
            conn.commit()
            flash("Cập nhật nội quy thành công!", "success")
            return redirect(url_for('noi_quy'))
        except Exception as e:
            flash(f"Lỗi khi cập nhật: {e}", "error")

    cursor.close()
    conn.close()
    return render_template('edit_noi_quy.html', rule=rule)

# Thêm dữ liệu vi phạm nội quy


@app.route('/add_noi_quy', methods=['GET', 'POST'])
@login_required
def add_noi_quy():
    if session.get('role') not in ['admin', 'giamthi', 'user']:
        flash("Bạn không có quyền thêm dữ liệu vi phạm nội quy.", 'error')
        return redirect(url_for('noi_quy'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Lấy danh sách tuần & lớp trực
    cursor.execute("SELECT DISTINCT tuan FROM phan_cong_truc ORDER BY tuan")
    available_tuan = [row['tuan'] for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT lop_truc FROM phan_cong_truc ORDER BY lop_truc")
    available_lop_truc = [row['lop_truc'] for row in cursor.fetchall()]

    user_tuan = session.get('tuan', '')
    user_lop = session.get('lop', '')

    if request.method == 'POST':
        # Lấy dữ liệu an toàn, tránh NoneType.strip
        tuan = (request.form.get('tuan') or user_tuan or '').strip()
        lop_truc = (request.form.get('lop_truc') or session.get('lop_truc', '') or '').strip()

        if session.get('role') == 'admin':
            lop = (request.form.get('lop') or '').strip()
        elif session.get('role') == 'giamthi':
            lop = (request.form.get('lop') or user_lop or '').strip()
        else:  # user
            lop = user_lop or ''

        noi_dung_vi_pham = (request.form.get('vi_pham') or '').strip()
        diem_tru = int(request.form.get('diem_tru') or 0)
        so_luot = int(request.form.get('so_luot') or 0)
        hoc_sinh = (request.form.get('hoc_sinh') or '').strip()
        tong_diem = diem_tru * so_luot

        try:
            cursor.execute("""
                INSERT INTO rules_data 
                (tuan, lop, lop_truc, noi_dung_vi_pham, diem_tru, tong_diem_vi_pham, so_luot_vi_pham, ten_hoc_sinh_vi_pham)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (tuan, lop, lop_truc, noi_dung_vi_pham, diem_tru, tong_diem, so_luot, hoc_sinh))
            conn.commit()
            flash("Đã thêm dữ liệu vi phạm nội quy thành công.", 'success')
        except mysql.connector.Error as err:
            flash(f"Lỗi khi thêm dữ liệu: {err}", 'error')
        finally:
            cursor.close()
            conn.close()

        return redirect(url_for('noi_quy'))

    cursor.close()
    conn.close()

    return render_template(
        'add_noi_quy.html',
        role=session.get('role'),
        user_tuan=user_tuan,
        available_tuan=available_tuan,
        available_lop_truc=available_lop_truc,
        lop_truc=session.get('lop_truc', '')
    )
@app.route('/download_account_template')
@login_required
def download_account_template():
    if session.get('role') != 'admin':
        flash("Bạn không có quyền tải mẫu.", 'error')
        return redirect(url_for('index'))

    wb = Workbook()
    ws = wb.active
    ws.append(["Tên", "Tên Người Dùng", "Lớp", "Cấp Quản Lý"])  # Không có mật khẩu

    file_path = os.path.join('uploads', 'account_template.xlsx')
    os.makedirs('uploads', exist_ok=True)
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)



# @app.route('/tong_ket', methods=['GET', 'POST'])
# @login_required
# def tong_ket():
#     if session.get('role') not in ['admin', 'user', 'viewer']:
#         flash("Bạn không có quyền truy cập trang tổng kết.", 'error')
#         return redirect(url_for('home'))
#     conn = get_db_connection()
#     cursor = conn.cursor(dictionary=True)
#     user_lop_truc = session.get('lop_truc')
#     user_tuan = session.get('tuan')
    
#     # Kiểm tra trạng thái tổng kết
#     trangthai_tongket = 'Chưa tổng kết'
#     if user_lop_truc and user_tuan:
#         cursor.execute("SELECT trangthai FROM bang_tong_ket WHERE lop_truc = %s AND tuan = %s", (user_lop_truc, user_tuan))
#         result = cursor.fetchone()
#         cursor.fetchall()  # Đọc hết kết quả
#         if result:
#             trangthai_tongket = result['trangthai']
    
#     if request.method == 'POST':
#         if 'recalculate' in request.form:
#             if session.get('role') == 'user' and trangthai_tongket == 'Đã tổng kết':
#                 flash(f"Tuần {user_tuan} đã được tổng kết cho lớp trực {user_lop_truc}.", 'warning')
#                 cursor.close()
#                 conn.close()
#                 return redirect(url_for('tong_ket'))
#             if session.get('role') == 'user' and (not user_lop_truc or not user_tuan):
#                 flash("Bạn cần được gán lớp trực và tuần trước khi tổng kết.", 'error')
#                 cursor.close()
#                 conn.close()
#                 return redirect(url_for('home'))
            
#             # Tạo con trỏ mới để tránh xung đột
#             cursor_temp = conn.cursor(dictionary=True)
#             if session.get('role') == 'admin':
#                 cursor_temp.execute("SELECT DISTINCT tuan, lop_truc FROM study_data WHERE lop_truc IS NOT NULL UNION SELECT DISTINCT tuan, lop_truc FROM rules_data WHERE lop_truc IS NOT NULL")
#             else:
#                 cursor_temp.execute("SELECT DISTINCT tuan, lop_truc FROM study_data WHERE lop_truc = %s AND tuan = %s AND lop_truc IS NOT NULL UNION SELECT DISTINCT tuan, lop_truc FROM rules_data WHERE lop_truc = %s AND tuan = %s AND lop_truc IS NOT NULL", (user_lop_truc, user_tuan, user_lop_truc, user_tuan))
#             unique_weeks_lops = cursor_temp.fetchall()  # Đọc hết kết quả
#             cursor_temp.close()  # Đóng con trỏ tạm
            
#             if not unique_weeks_lops:
#                 flash("Không có dữ liệu hợp lệ trong study_data hoặc rules_data để tổng kết.", 'warning')
#                 cursor.close()
#                 conn.close()
#                 return redirect(url_for('tong_ket'))
            
#             if session.get('role') == 'admin':
#                 cursor.execute("TRUNCATE TABLE bang_tong_ket")
#                 conn.commit()
            
#             for entry in unique_weeks_lops:
#                 tuan = entry['tuan']
#                 lop_truc = entry['lop_truc']
#                 if not lop_truc:
#                     continue
#                 cursor.execute("SELECT khoi FROM phan_cong_truc WHERE lop_truc = %s AND tuan = %s", (lop_truc, tuan))
#                 khoi_result = cursor.fetchone()
#                 cursor.fetchall()  # Đọc hết kết quả
#                 khoi = khoi_result['khoi'] if khoi_result else 'Unknown'
#                 cursor.execute("SELECT SUM(tong_diem) as total_study_points FROM study_data WHERE tuan = %s AND lop_truc = %s", (tuan, lop_truc))
#                 study_result = cursor.fetchone()
#                 cursor.fetchall()  # Đọc hết kết quả
#                 total_study_points = study_result['total_study_points'] if study_result and study_result['total_study_points'] is not None else 0
#                 cursor.execute("SELECT SUM(tong_diem_vi_pham) as total_rules_points FROM rules_data WHERE tuan = %s AND lop_truc = %s", (tuan, lop_truc))
#                 rules_result = cursor.fetchone()
#                 cursor.fetchall()  # Đọc hết kết quả
#                 total_rules_points = rules_result['total_rules_points'] if rules_result and rules_result['total_rules_points'] is not None else 0
#                 tong_diem_chung = total_study_points + total_rules_points
#                 cursor.execute("""
#                     INSERT INTO bang_tong_ket (tuan, khoi, lop_truc, tong_diem_hoc_tap, tong_diem_noi_quy, tong_diem_chung, trangthai)
#                     VALUES (%s, %s, %s, %s, %s, %s, %s)
#                     ON DUPLICATE KEY UPDATE 
#                         khoi = %s,
#                         tong_diem_hoc_tap = %s, 
#                         tong_diem_noi_quy = %s, 
#                         tong_diem_chung = %s,
#                         trangthai = %s
#                 """, (
#                     tuan, khoi, lop_truc, total_study_points, total_rules_points, tong_diem_chung, 'Đã tổng kết',
#                     khoi, total_study_points, total_rules_points, tong_diem_chung, 'Đã tổng kết'
#                 ))
#                 cursor.fetchall()  # Đọc hết kết quả
#                 if session.get('role') == 'user':
#                     cursor.execute("UPDATE accounts SET trangthai = 'Đã tổng kết' WHERE username = %s AND tuan = %s", (session.get('username'), user_tuan))
#                     cursor.fetchall()  # Đọc hết kết quả
#             conn.commit()
#             flash(f"Đã tổng kết thành công cho {'tất cả lớp trực' if session.get('role') == 'admin' else f'lớp trực {user_lop_truc}'} – Tuần {user_tuan}.", 'success')
#             cursor.close()
#             conn.close()
#             return redirect(url_for('tong_ket'))
        
#         elif 'delete_class' in request.form and session.get('role') == 'admin':
#             lop_truc = request.form.get('lop_truc')
#             tuan = request.form.get('tuan')
#             try:
#                 cursor.execute("DELETE FROM bang_tong_ket WHERE lop_truc = %s AND tuan = %s", (lop_truc, tuan))
#                 conn.commit()
#                 flash(f"Đã xóa dữ liệu tổng kết của lớp {lop_truc} cho tuần {tuan}.", 'success')
#             except mysql.connector.Error as err:
#                 flash(f"Lỗi khi xóa dữ liệu tổng kết: {err}", 'error')
#             cursor.close()
#             conn.close()
#             return redirect(url_for('tong_ket'))
    
#     selected_tuan = request.args.get('tuan', type=str)
#     selected_khoi = request.args.get('khoi', type=str)
#     view_all = request.args.get('view_all', type=str) == 'true'
    
#     cursor.execute("SELECT DISTINCT tuan FROM bang_tong_ket ORDER BY tuan ASC")
#     available_weeks = [row['tuan'] for row in cursor.fetchall()]
#     cursor.execute("SELECT DISTINCT khoi FROM bang_tong_ket WHERE khoi IS NOT NULL ORDER BY khoi ASC")
#     available_khoi = [row['khoi'] for row in cursor.fetchall()]
    
#     query = "SELECT tuan, khoi, lop_truc, tong_diem_hoc_tap, tong_diem_noi_quy, tong_diem_chung, trangthai FROM bang_tong_ket WHERE 1=1"
#     params = []
#     if selected_tuan and not view_all:
#         query += " AND tuan = %s"
#         params.append(selected_tuan)
#     if selected_khoi:
#         query += " AND khoi = %s"
#         params.append(selected_khoi)
#     if session.get('role') == 'user':
#         if user_lop_truc:
#             query += " AND lop_truc = %s"
#             params.append(user_lop_truc)
#     query += " ORDER BY tuan ASC, khoi ASC, tong_diem_chung DESC"
#     cursor.execute(query, params)
#     data = cursor.fetchall()
#     cursor.fetchall()  # Đọc hết kết quả
    
#     if not data:
#         flash("Không có dữ liệu tổng kết để hiển thị.", 'warning')
#         cursor.close()
#         conn.close()
#         return render_template('tong_ket.html', data=[], available_weeks=available_weeks, available_khoi=available_khoi, selected_tuan=selected_tuan, selected_khoi=selected_khoi, trangthai_tongket=trangthai_tongket)
    
#     ranked_data = []
#     grouped_data = defaultdict(list)
#     for item in data:
#         grouped_data[(item['tuan'], item.get('khoi', 'Unknown'))].append(item)
#     for tuan_khoi in sorted(grouped_data.keys()):
#         current_week_block_data = sorted(grouped_data[tuan_khoi], key=lambda x: x['tong_diem_chung'], reverse=True)
#         current_rank = 1
#         same_rank_count = 1
#         prev_diem = None
#         for i, item in enumerate(current_week_block_data):
#             if i == 0:
#                 item['xep_hang'] = current_rank
#             else:
#                 if item['tong_diem_chung'] < prev_diem:
#                     current_rank += same_rank_count
#                     same_rank_count = 1
#                 else:
#                     same_rank_count += 1
#                 item['xep_hang'] = current_rank
#             prev_diem = item['tong_diem_chung']
#             ranked_data.append(item)
    
#     cursor.close()
#     conn.close()
#     return render_template('tong_ket.html', data=ranked_data, available_weeks=available_weeks, available_khoi=available_khoi, selected_tuan=selected_tuan, selected_khoi=selected_khoi, trangthai_tongket=trangthai_tongket)

@app.route('/api/tuan_loptruc')
@login_required
def api_tuan_loptruc():
    if session.get('role') not in ['admin', 'giamthi']:
        return jsonify({"error": "Không có quyền"}), 403

    import requests
    url = "https://sheetdb.io/api/v1/bhx4aotm9houf"  # API nguồn dữ liệu
    res = requests.get(url)
    data = res.json()

    tuan_list = sorted(set(row.get("Tuan", "") for row in data if row.get("Tuan")))
    lop_list = sorted(set(row.get("Lop", "") for row in data if row.get("Lop")))

    return jsonify({"tuan": tuan_list, "lop": lop_list})


@app.route('/export_accounts', methods=['GET'])
@login_required
def export_accounts():
    if session.get('role') != 'admin':
        flash("Bạn không có quyền xuất danh sách tài khoản.", 'error')
        return redirect(url_for('index'))

    selected_tuan = request.args.get('tuan', type=int)

    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        sql = """
        SELECT Name, username, password, role, lop, tuan, Capquanli, lop_truc, trangthai
        FROM accounts
        WHERE role = 'user'
        """
        params = []

        if selected_tuan:
            sql += " AND tuan = %s"
            params.append(selected_tuan)

        cursor.execute(sql, tuple(params))
        data = cursor.fetchall()

        if not data:
            flash("Không có dữ liệu tài khoản user để xuất.", 'warning')
            return redirect(url_for('index'))

        df = pd.DataFrame(data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Accounts_User')

        output.seek(0)
        return send_file(output,
                         download_name=f"danh_sach_user_tuan_{selected_tuan or 'all'}.xlsx",
                         as_attachment=True)

    except Exception as e:
        flash(f"Lỗi khi xuất danh sách: {e}", 'error')
        return redirect(url_for('index'))
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/tong_ket', methods=['GET', 'POST'])
@login_required
def tong_ket():
    if session.get('role') not in ['admin', 'user', 'viewer', 'giamthi']:
        flash("Bạn không có quyền truy cập trang tổng kết.", 'error')
        return redirect(url_for('home'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    user_lop_truc = session.get('lop_truc')
    user_tuan = session.get('tuan')

    # Kiểm tra trạng thái tổng kết của user
    trangthai_tongket = 'Chưa tổng kết'
    if user_lop_truc and user_tuan:
        cursor.execute(
            "SELECT trangthai FROM bang_tong_ket WHERE lop_truc = %s AND tuan = %s",
            (user_lop_truc, user_tuan)
        )
        result = cursor.fetchone()
        cursor.fetchall()
        if result:
            trangthai_tongket = result['trangthai']

    # Lấy danh sách lớp chưa tổng kết (cho admin/giamthi)
    cursor.execute("""
        SELECT DISTINCT p.tuan, p.lop_truc
        FROM phan_cong_truc p
        LEFT JOIN bang_tong_ket b
            ON p.tuan = b.tuan AND p.lop_truc = b.lop_truc
        WHERE b.lop_truc IS NULL
        ORDER BY p.tuan, p.lop_truc
    """)
    not_summarized = cursor.fetchall()
    available_weeks = sorted(set(row['tuan'] for row in not_summarized))
    available_lop_truc = sorted(set(row['lop_truc'] for row in not_summarized))

    def tong_ket_lop(lop_truc, tuan):
        """Hàm tính toán & lưu tổng kết cho 1 lớp."""
        cursor.execute(
            "SELECT khoi FROM phan_cong_truc WHERE lop_truc = %s AND tuan = %s",
            (lop_truc, tuan)
        )
        khoi_result = cursor.fetchone()
        cursor.fetchall()
        khoi = khoi_result['khoi'] if khoi_result else 'Unknown'

        cursor.execute(
            "SELECT SUM(tong_diem) as total_study_points FROM study_data WHERE tuan = %s AND lop_truc = %s",
            (tuan, lop_truc)
        )
        study_result = cursor.fetchone()
        cursor.fetchall()
        total_study_points = study_result['total_study_points'] if study_result and study_result['total_study_points'] is not None else 0

        cursor.execute(
            "SELECT SUM(tong_diem_vi_pham) as total_rules_points FROM rules_data WHERE tuan = %s AND lop_truc = %s",
            (tuan, lop_truc)
        )
        rules_result = cursor.fetchone()
        cursor.fetchall()
        total_rules_points = rules_result['total_rules_points'] if rules_result and rules_result['total_rules_points'] is not None else 0

        tong_diem_chung = total_study_points + total_rules_points

        cursor.execute("""
            INSERT INTO bang_tong_ket (tuan, khoi, lop_truc, tong_diem_hoc_tap, tong_diem_noi_quy, tong_diem_chung, trangthai)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE 
                khoi = %s,
                tong_diem_hoc_tap = %s, 
                tong_diem_noi_quy = %s, 
                tong_diem_chung = %s,
                trangthai = %s
        """, (
            tuan, khoi, lop_truc, total_study_points, total_rules_points, tong_diem_chung, 'Đã tổng kết',
            khoi, total_study_points, total_rules_points, tong_diem_chung, 'Đã tổng kết'
        ))

    # Xử lý POST
    if request.method == 'POST':
        # Tổng kết 1 lớp
        if 'recalculate' in request.form:
            if session.get('role') == 'user':
                if trangthai_tongket == 'Đã tổng kết':
                    flash(f"Tuần {user_tuan} đã được tổng kết cho lớp trực {user_lop_truc}.", 'warning')
                elif not user_lop_truc or not user_tuan:
                    flash("Bạn cần được gán lớp trực và tuần trước khi tổng kết.", 'error')
                else:
                    tong_ket_lop(user_lop_truc, user_tuan)
                    conn.commit()
                    flash(f"Đã tổng kết thành công cho lớp trực {user_lop_truc} – Tuần {user_tuan}.", 'success')
            elif session.get('role') in ['admin', 'giamthi']:
                lop_truc = request.form.get('lop_truc')
                tuan = request.form.get('tuan')
                if not lop_truc or not tuan:
                    flash("Vui lòng chọn lớp trực và tuần để tổng kết.", 'error')
                else:
                    tong_ket_lop(lop_truc, tuan)
                    conn.commit()
                    flash(f"Đã tổng kết thành công cho lớp trực {lop_truc} – Tuần {tuan}.", 'success')

        # Tổng kết tất cả lớp chưa tổng kết
        elif 'recalculate_all' in request.form and session.get('role') in ['admin', 'giamthi']:
            for row in not_summarized:
                tong_ket_lop(row['lop_truc'], row['tuan'])
            conn.commit()
            flash(f"Đã tổng kết tất cả {len(not_summarized)} lớp chưa tổng kết.", 'success')

        # Xóa dữ liệu tổng kết
        elif 'delete_class' in request.form and session.get('role') in ['admin', 'giamthi']:
            lop_truc = request.form.get('lop_truc')
            tuan = request.form.get('tuan')
            try:
                cursor.execute("DELETE FROM bang_tong_ket WHERE lop_truc = %s AND tuan = %s", (lop_truc, tuan))
                conn.commit()
                flash(f"Đã xóa dữ liệu tổng kết của lớp {lop_truc} cho tuần {tuan}.", 'success')
            except mysql.connector.Error as err:
                flash(f"Lỗi khi xóa dữ liệu tổng kết: {err}", 'error')

        cursor.close()
        conn.close()
        return redirect(url_for('tong_ket'))

    # Dữ liệu hiển thị bảng xếp hạng
    selected_tuan = request.args.get('tuan', type=str)
    selected_khoi = request.args.get('khoi', type=str)
    view_all = request.args.get('view_all', type=str) == 'true'

    cursor.execute("SELECT DISTINCT tuan FROM bang_tong_ket ORDER BY tuan ASC")
    available_weeks_all = [row['tuan'] for row in cursor.fetchall()]
    cursor.execute("SELECT DISTINCT khoi FROM bang_tong_ket WHERE khoi IS NOT NULL ORDER BY khoi ASC")
    available_khoi = [row['khoi'] for row in cursor.fetchall()]

    query = "SELECT tuan, khoi, lop_truc, tong_diem_hoc_tap, tong_diem_noi_quy, tong_diem_chung, trangthai FROM bang_tong_ket WHERE 1=1"
    params = []
    if selected_tuan and not view_all:
        query += " AND tuan = %s"
        params.append(selected_tuan)
    if selected_khoi:
        query += " AND khoi = %s"
        params.append(selected_khoi)
    query += " ORDER BY tuan ASC, khoi ASC, tong_diem_chung DESC"
    cursor.execute(query, params)
    data = cursor.fetchall()

    # Xếp hạng
    ranked_data = []
    grouped_data = defaultdict(list)
    for item in data:
        grouped_data[(item['tuan'], item.get('khoi', 'Unknown'))].append(item)
    for tuan_khoi in sorted(grouped_data.keys()):
        current_week_block_data = sorted(grouped_data[tuan_khoi], key=lambda x: x['tong_diem_chung'], reverse=True)
        current_rank = 1
        same_rank_count = 1
        prev_diem = None
        for i, item in enumerate(current_week_block_data):
            if i == 0:
                item['xep_hang'] = current_rank
            else:
                if item['tong_diem_chung'] < prev_diem:
                    current_rank += same_rank_count
                    same_rank_count = 1
                else:
                    same_rank_count += 1
                item['xep_hang'] = current_rank
            prev_diem = item['tong_diem_chung']
            ranked_data.append(item)

    cursor.close()
    conn.close()
    return render_template(
        'tong_ket.html',
        data=ranked_data,
        available_weeks=available_weeks,
        available_lop_truc=available_lop_truc,
        available_khoi=available_khoi,
        selected_tuan=selected_tuan,
        selected_khoi=selected_khoi,
        trangthai_tongket=trangthai_tongket
    )



@app.route('/update_study_data_admin/<int:data_id>', methods=['GET', 'POST'])
@login_required
def update_study_data_admin(data_id):
    # Chỉ cho admin và giamthi truy cập
    if session.get('role') not in ['admin', 'giamthi']:
        flash("Bạn không có quyền chỉnh sửa dữ liệu học tập.", 'error')
        return redirect(url_for('hoc_tap'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Lấy dữ liệu để kiểm tra tồn tại
    cursor.execute("SELECT tuan, lop_truc FROM study_data WHERE id = %s", (data_id,))
    study_data = cursor.fetchone()
    cursor.fetchall()
    if not study_data:
        flash("Không tìm thấy dữ liệu học tập.", 'error')
        cursor.close()
        conn.close()
        return redirect(url_for('hoc_tap'))

    # Kiểm tra trạng thái tổng kết
    cursor.execute("""
        SELECT trangthai FROM bang_tong_ket 
        WHERE lop_truc = %s AND tuan = %s
    """, (study_data['lop_truc'], study_data['tuan']))
    account_status = cursor.fetchone()
    cursor.fetchall()
    if account_status and account_status['trangthai'] == 'Đã tổng kết':
        flash("Không thể chỉnh sửa dữ liệu học tập vì tuần này đã được tổng kết.", 'warning')
        cursor.close()
        conn.close()
        return redirect(url_for('hoc_tap'))

    # Cập nhật dữ liệu
    if request.method == 'POST':
        gio_a = int(request.form.get('gio_a') or 0)
        gio_b = int(request.form.get('gio_b') or 0)
        gio_c = int(request.form.get('gio_c') or 0)
        gio_d = int(request.form.get('gio_d') or 0)
        dat_kieu_mau = request.form['dat_kieu_mau']

        tong_diem = gio_a * 5 + gio_b * -5 + gio_c * -15 + gio_d * -25
        tong_diem += 5 if dat_kieu_mau == "Yes" else -10

        try:
            cursor.execute("""
                UPDATE study_data 
                SET gio_a = %s, gio_b = %s, gio_c = %s, gio_d = %s, 
                    dat_kieu_mau = %s, tong_diem = %s
                WHERE id = %s
            """, (gio_a, gio_b, gio_c, gio_d, dat_kieu_mau, tong_diem, data_id))
            conn.commit()
            flash("Đã cập nhật dữ liệu học tập thành công.", 'success')
        except mysql.connector.Error as err:
            flash(f"Lỗi khi cập nhật dữ liệu học tập: {err}", 'error')
        finally:
            cursor.close()
            conn.close()
        return redirect(url_for('hoc_tap'))

    # Lấy dữ liệu để hiển thị form
    cursor.execute("SELECT * FROM study_data WHERE id = %s", (data_id,))
    data = cursor.fetchone()
    cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template('update_study_data_admin.html', data=data)




@app.route('/export_user_tong_ket', methods=['GET'])
@login_required
def export_user_tong_ket():
    if session.get('role') not in ['admin', 'user', 'viewer', 'giamthi']:
        flash("Bạn không có quyền xuất dữ liệu tổng kết.", 'error')
        return redirect(url_for('home'))
    
    selected_tuan = request.args.get('tuan', type=str)
    view_all = request.args.get('view_all', type=str) == 'true'
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    query = "SELECT tuan, khoi, lop_truc, tong_diem_hoc_tap, tong_diem_noi_quy, tong_diem_chung FROM bang_tong_ket WHERE 1=1"
    params = []
    if session.get('role') == 'user' and not view_all:
        user_lop_truc = session.get('lop_truc')
        user_tuan = session.get('tuan')
        if user_lop_truc and user_tuan:
            query += " AND lop_truc = %s AND tuan = %s"
            params.extend([user_lop_truc, user_tuan])
    elif selected_tuan and not view_all:
        query += " AND tuan = %s"
        params.append(selected_tuan)
    query += " ORDER BY tuan ASC, khoi ASC, tong_diem_chung DESC"
    
    cursor.execute(query, params)
    data = cursor.fetchall()
    cursor.fetchall()  # Đọc hết kết quả
    cursor.close()
    conn.close()
    
    if not data:
        flash("Không có dữ liệu tổng kết để xuất.", 'warning')
        return redirect(url_for('tong_ket'))
    
    # Tính toán xếp hạng
    ranked_data = []
    grouped_data = defaultdict(list)
    for item in data:
        grouped_data[(item['tuan'], item.get('khoi', 'Unknown'))].append(item)
    for tuan_khoi in sorted(grouped_data.keys()):
        current_week_block_data = sorted(grouped_data[tuan_khoi], key=lambda x: x['tong_diem_chung'], reverse=True)
        current_rank = 1
        same_rank_count = 1
        prev_diem = None
        for i, item in enumerate(current_week_block_data):
            if i == 0:
                item['xep_hang'] = current_rank
            else:
                if item['tong_diem_chung'] < prev_diem:
                    current_rank += same_rank_count
                    same_rank_count = 1
                else:
                    same_rank_count += 1
                item['xep_hang'] = current_rank
            prev_diem = item['tong_diem_chung']
            ranked_data.append(item)
    
    # Tạo DataFrame với tên cột tiếng Việt
    df = pd.DataFrame(ranked_data, columns=['tuan', 'khoi', 'lop_truc', 'tong_diem_hoc_tap', 'tong_diem_noi_quy', 'tong_diem_chung', 'xep_hang'])
    df = df.rename(columns={
        'tuan': 'Tuần',
        'khoi': 'Khối',
        'lop_truc': 'Lớp',
        'tong_diem_hoc_tap': 'Tổng Điểm Học Tập',
        'tong_diem_noi_quy': 'Tổng Điểm Nội Quy',
        'tong_diem_chung': 'Tổng Điểm Chung',
        'xep_hang': 'Xếp Hạng'
    })
    
    # Xuất Excel với định dạng thẩm mỹ
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Bảng Tổng Kết', index=False, startrow=2)
        workbook = writer.book
        worksheet = writer.sheets['Bảng Tổng Kết']
        
        # Định dạng tiêu đề chính
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'font_color': '#1e40af',
            'align': 'center',
            'valign': 'vcenter'
        })
        worksheet.merge_range('A1:G1', 'BẢNG TỔNG KẾT THI ĐUA', title_format)
        
        # Định dạng tiêu đề cột
        header_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#1e40af',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial'
        })
        
        # Định dạng dữ liệu
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'left',
            'valign': 'vcenter',
            'font_name': 'Arial'
        })
        number_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': '0',
            'font_name': 'Arial'
        })
        
        # Áp dụng định dạng tiêu đề cột
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(2, col_num, value, header_format)
        
        # Áp dụng định dạng cho dữ liệu
        for row_num in range(3, len(df) + 3):
            for col_num in range(len(df.columns)):
                value = df.iloc[row_num - 3, col_num]
                if col_num in [3, 4, 5, 6]:  # Cột số: Tổng Điểm Học Tập, Nội Quy, Chung, Xếp Hạng
                    worksheet.write(row_num, col_num, value if value is not None else 0, number_format)
                else:
                    worksheet.write(row_num, col_num, value if value is not None else '', cell_format)
        
        # Tự động điều chỉnh độ rộng cột
        for col_num, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
            worksheet.set_column(col_num, col_num, max_len)
    
    output.seek(0)
    filename = f"tong_ket_{selected_tuan or 'tat_ca'}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # Tạo DataFrame với tên cột tiếng Việt
    df = pd.DataFrame(ranked_data, columns=['tuan', 'khoi', 'lop_truc', 'tong_diem_hoc_tap', 'tong_diem_noi_quy', 'tong_diem_chung', 'xep_hang'])
    df = df.rename(columns={
        'tuan': 'Tuần',
        'khoi': 'Khối',
        'lop_truc': 'Lớp',
        'tong_diem_hoc_tap': 'Tổng Điểm Học Tập',
        'tong_diem_noi_quy': 'Tổng Điểm Nội Quy',
        'tong_diem_chung': 'Tổng Điểm Chung',
        'xep_hang': 'Xếp Hạng'
    })
    
    # Xuất Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Tổng Kết', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Tổng Kết']
        for idx, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 1
            worksheet.set_column(idx, idx, max_len)
    output.seek(0)
    filename = f"tong_ket_tuan_{selected_tuan or 'tat_ca'}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export_summary', methods=['GET'])
@login_required
def export_summary():
    if session.get('role') not in ['admin', 'giamthi']:
        flash("Bạn không có quyền xuất dữ liệu này.", 'error')
        return redirect(url_for('home'))

    selected_tuan = request.args.get('export_tuan', type=str)
    selected_lop = request.args.get('export_lop', type=str)

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SET SESSION group_concat_max_len = 10000;")

        query = """
            SELECT 
                btk.tuan,
                btk.lop,
                btk.tong_diem_hoc_tap,
                btk.tong_diem_noi_quy,
                btk.tong_diem_chung,
                GROUP_CONCAT(DISTINCT rd.ten_hoc_sinh_vi_pham SEPARATOR '; ') AS ten_hoc_sinh_vi_pham,
                GROUP_CONCAT(DISTINCT CONCAT(rd.noi_dung_vi_pham, ' (', rd.diem_tru, ' điểm, ', rd.so_luot_vi_pham, ' lượt)') SEPARATOR '; ') AS chi_tiet_vi_pham
            FROM 
                bang_tong_ket btk
            LEFT JOIN 
                rules_data rd ON btk.lop = rd.lop AND btk.tuan = rd.tuan
        """
        params = []
        conditions = []

        if selected_tuan:
            conditions.append("btk.tuan = %s")
            params.append(selected_tuan)
        if selected_lop:
            conditions.append("btk.lop = %s")
            params.append(selected_lop)

        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        query += " GROUP BY btk.tuan, btk.lop ORDER BY btk.tuan, btk.lop"

        cursor.execute(query, tuple(params))
        summary_data = cursor.fetchall()

        if not summary_data:
            flash("Không tìm thấy dữ liệu tổng kết theo tuần và lớp đã chọn để xuất báo cáo.", 'warning')
            return redirect(url_for('home'))

        print(f"Summary data: {summary_data}")  # Debug
        df = pd.DataFrame(summary_data)
        
        if selected_tuan:
            df['xep_hang'] = df['tong_diem_chung'].rank(method='min', ascending=False).astype(int)
        else:
            df['xep_hang'] = df.groupby('tuan')['tong_diem_chung'].rank(method='min', ascending=False).astype(int)

        df = df[['tuan', 'lop', 'tong_diem_hoc_tap', 'tong_diem_noi_quy', 'tong_diem_chung', 'xep_hang', 'ten_hoc_sinh_vi_pham', 'chi_tiet_vi_pham']]
        df = df.rename(columns={
            'tuan': 'Tuần',
            'lop': 'Lớp',
            'tong_diem_hoc_tap': 'Tổng Điểm Học Tập',
            'tong_diem_noi_quy': 'Tổng Điểm Nội Quy',
            'tong_diem_chung': 'Tổng Điểm Chung',
            'xep_hang': 'Xếp Hạng',
            'ten_hoc_sinh_vi_pham': 'Tên Học Sinh Vi Phạm',
            'chi_tiet_vi_pham': 'Chi Tiết Vi Phạm Nội Quy'
        })

        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Tổng Kết Vi Phạm')
        writer.close()
        output.seek(0)

        filename = f"Tong_Ket_Vi_Pham_{selected_lop or 'TatCaLop'}_Tuan{selected_tuan or 'TatCaTuan'}.xlsx"
        return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except mysql.connector.Error as err:
        flash(f"Lỗi khi truy vấn dữ liệu từ cơ sở dữ liệu: {err}", 'error')
        print(f"Database Error: {err}")
        return redirect(url_for('home'))
    except Exception as e:
        flash(f"Lỗi không xác định khi xuất báo cáo Excel: {e}", 'error')
        print(f"General Error: {e}")
        return redirect(url_for('home'))
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

# Route xuất báo cáo học tập ra PDF
@app.route('/export_hoc_tap_pdf', methods=['GET'])
@login_required
def export_hoc_tap_pdf():
    if session.get('role') not in ['admin', 'user', 'viewer', 'giamthi']:
        flash("Bạn không có quyền xuất báo cáo này.", 'error')
        return redirect(url_for('hoc_tap'))

    selected_tuan = request.args.get('tuan', type=str)
    selected_lop = request.args.get('lop', type=str)

    if not selected_tuan or not selected_lop:
        flash("Vui lòng chọn Tuần và Lớp để xuất báo cáo PDF.", 'warning')
        return redirect(url_for('hoc_tap'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    query = "SELECT * FROM study_data WHERE tuan = %s AND lop = %s ORDER BY id ASC"
    cursor.execute(query, (selected_tuan, selected_lop))
    data = cursor.fetchall()
    cursor.close()
    conn.close()

    if not data:
        flash(f"Không có dữ liệu học tập cho Tuần {selected_tuan} và Lớp {selected_lop} để xuất báo cáo PDF.", 'warning')
        return redirect(url_for('hoc_tap', tuan=selected_tuan, lop=selected_lop))

    pdf = FPDF()
    pdf.add_page()
    pdf.add_font('DejaVuSans', '', 'DejaVuSansCondensed.ttf', uni=True)
    pdf.set_font('DejaVuSans', '', 12)

    pdf.cell(0, 10, f"BÁO CÁO HỌC TẬP - Tuần {selected_tuan} - Lớp {selected_lop}", 0, 1, 'C')
    pdf.ln(5)

    headers = ["Tuần", "Lớp", "Giờ A", "Giờ B", "Giờ C", "Giờ D", "Đạt Kiểu Mẫu", "Tổng Điểm"]
    col_widths = [15, 15, 15, 15, 15, 15, 30, 20]

    pdf.set_font('DejaVuSans', '', 10)
    pdf.set_fill_color(200, 220, 255)
    for i, header in enumerate(headers):
        if i < len(col_widths):
            pdf.cell(col_widths[i], 10, header, 1, 0, 'C', 1)
    pdf.ln()

    pdf.set_font('DejaVuSans', '', 9)
    for row in data:
        pdf.cell(col_widths[0], 10, str(row.get('tuan', '')), 1, 0, 'C')
        pdf.cell(col_widths[1], 10, str(row.get('lop', '')), 1, 0, 'C')
        pdf.cell(col_widths[2], 10, str(row.get('gio_a', 0)), 1, 0, 'C')
        pdf.cell(col_widths[3], 10, str(row.get('gio_b', 0)), 1, 0, 'C')
        pdf.cell(col_widths[4], 10, str(row.get('gio_c', 0)), 1, 0, 'C')
        pdf.cell(col_widths[5], 10, str(row.get('gio_d', 0)), 1, 0, 'C')
        pdf.cell(col_widths[6], 10, str(row.get('dat_kieu_mau', '')), 1, 0, 'C')
        pdf.cell(col_widths[7], 10, str(row.get('tong_diem', 0)), 1, 0, 'C')
        pdf.ln()

    pdf_output = pdf.output(dest='S').encode('latin-1')
    return send_file(BytesIO(pdf_output), mimetype='application/pdf', as_attachment=True, download_name=f"BaoCaoHocTap_Tuan_{selected_tuan}_Lop_{selected_lop}.pdf")

# Gán tuần cho lớp (admin sử dụng)
@app.route('/assign_tuan', methods=['GET', 'POST'])
@login_required
def assign_tuan():
    if session.get('role') != ['admin', 'giamthi']:
        flash("Bạn không có quyền phân công.", 'error')
        return redirect(url_for('home'))

    if request.method == 'POST':
        lop = request.form['lop'].strip()
        tuan = request.form['tuan'].strip()
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO week_assignment (lop, tuan)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE tuan=%s
        """, (lop, tuan, tuan))
        conn.commit()
        cursor.close()
        conn.close()
        flash(f"Đã gán tuần {tuan} cho lớp {lop} thành công.", 'success')
        return redirect(url_for('assign_tuan'))
    return render_template('assign_tuan.html')


@app.route('/index', methods=['GET'])
@login_required
def index():
    if session.get('role') != 'admin':
        flash("Bạn không có quyền truy cập trang quản lý tài khoản.", 'error')
        return redirect(url_for('home'))

    accounts = []
    conn = None
    selected_tuan = request.args.get('tuan', type=int)

    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        sql_query = "SELECT id, Name, username, password, role, lop, tuan, Capquanli, lop_truc, trangthai FROM accounts"
        query_params = []

        if selected_tuan is not None:
            sql_query += " WHERE tuan = %s"
            query_params.append(selected_tuan)
        
        sql_query += " ORDER BY id DESC"

        cursor.execute(sql_query, tuple(query_params))
        accounts = cursor.fetchall()

    except mysql.connector.Error as err:
        print(f"Lỗi khi đọc dữ liệu từ DB: {err}")
        flash(f"Lỗi khi tải danh sách tài khoản: {err}", 'error')
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

    return render_template('insert_account.html', accounts=accounts, selected_tuan=selected_tuan)

def get_db():
    conn = mysql.connector.connect(**db_config)
    return conn

def create_table():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS phan_cong_truc (
            id INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
            khoi VARCHAR(10) NOT NULL,
            lop VARCHAR(10) NOT NULL,
            tuan INT NOT NULL,
            lop_truc VARCHAR(10) NOT NULL
        )
    ''')
    conn.commit()
    cursor.close()
    conn.close()

def save_phancong(khoi, lop, phan_cong):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM phan_cong_truc WHERE khoi=%s AND lop=%s", (khoi, lop))
    for tuan, lop_truc in enumerate(phan_cong, start=1):
        cursor.execute("INSERT INTO phan_cong_truc (khoi, lop, tuan, lop_truc) VALUES (%s, %s, %s, %s)",
                       (khoi, lop, tuan, lop_truc))
    conn.commit()
    cursor.close()
    conn.close()
@app.route('/insert_account', methods=['GET'])
@login_required
def insert_account_form():
    if 'username' not in session or session['role'] != 'admin':
        flash("Bạn không có quyền truy cập trang này.", 'error')
        return redirect(url_for('login'))
    
    return render_template('insert_account.html')

@app.route('/add_account', methods=['POST'])
@login_required
def add_account():
    if 'username' not in session or session['role'] != 'admin':
        flash("Bạn không có quyền thực hiện hành động này.", 'error')
        return redirect(url_for('login'))

    input_tuan = request.form.get('current_tuan_for_add', type=int) or 1
    
    input_name = request.form['name']
    input_username = request.form['username']
    input_lop = request.form['lop']
    input_capquanli = request.form['Capquanli']
    
    input_password = generate_specific_password()
    input_role = input_capquanli

    conn = None
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        sql = """
        INSERT INTO accounts (Name, username, password, role, lop, tuan, Capquanli, trangthai)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(sql, (input_name, input_username, input_password,
                              input_role, input_lop, input_tuan, input_capquanli, 'Chưa tổng kết'))

        conn.commit()
        flash(f"Thêm tài khoản thành công! Tên người dùng: **{input_username}**, Mật khẩu: **{input_password}**, Tuần: **{input_tuan}**", 'success')

    except mysql.connector.Error as err:
        flash(f"Lỗi khi thêm tài khoản: {err}", 'error')
        print(f"Error: {err}")

    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

    return redirect(url_for('index', tuan=input_tuan if input_tuan != 1 else None))




@app.route('/toggle_status/<int:account_id>', methods=['POST'])
@login_required
def toggle_status(account_id):
    if session.get('role') != 'admin':
        flash("Bạn không có quyền thay đổi trạng thái tài khoản.", 'error')
        return redirect(url_for('index'))

    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        cursor.execute("SELECT trangthai FROM accounts WHERE id = %s", (account_id,))
        result = cursor.fetchone()
        if result:
            current_status = result[0]
            new_status = 'Đã tổng kết' if current_status == 'Đã tổng kết' else 'Đã tổng kết'
            cursor.execute("UPDATE accounts SET trangthai = %s WHERE id = %s", (new_status, account_id))
            conn.commit()
            flash("Trạng thái đã được cập nhật!", "success")
        else:
            flash("Không tìm thấy tài khoản.", "error")

    except mysql.connector.Error as err:
        flash(f"Lỗi cơ sở dữ liệu: {err}", 'error')
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

    return redirect(url_for('index'))

@app.route('/edit_account/<int:account_id>', methods=['GET', 'POST'])
@login_required
def edit_account(account_id):
    if session.get('role') != 'admin':
        flash("Bạn không có quyền chỉnh sửa tài khoản.", 'error')
        return redirect(url_for('index'))

    conn = None
    account_to_edit = None
    accounts = []
    selected_tuan = request.args.get('tuan', type=int)

    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        sql_query = "SELECT id, Name, username, password, role, lop, tuan, Capquanli, lop_truc, trangthai FROM accounts"
        query_params = []
        if selected_tuan is not None:
            sql_query += " WHERE tuan = %s"
            query_params.append(selected_tuan)
        sql_query += " ORDER BY id DESC"

        cursor.execute(sql_query, tuple(query_params))
        accounts = cursor.fetchall()

        if request.method == 'POST':
            input_name = request.form['name']
            input_username = request.form['username']
            input_lop = request.form['lop']
            input_capquanli = request.form['Capquanli']
            input_tuan_edit = request.form.get('tuan_edit', type=int)
            input_lop_truc = request.form.get('lop_truc', '')
            input_trangthai = request.form.get('trangthai', 'Chưa tổng kết')

            cursor.execute("SELECT password FROM accounts WHERE id = %s", (account_id,))
            old_password_row = cursor.fetchone()
            if old_password_row:
                input_password = old_password_row['password']
            else:
                flash("Không tìm thấy tài khoản để cập nhật.", 'error')
                return redirect(url_for('index', tuan=selected_tuan))

            input_role = input_capquanli

            sql = """
            UPDATE accounts
            SET Name = %s, username = %s, password = %s, role = %s, lop = %s, tuan = %s, Capquanli = %s, lop_truc = %s, trangthai = %s
            WHERE id = %s
            """
            cursor.execute(sql, (input_name, input_username, input_password,
                                 input_role, input_lop, input_tuan_edit, input_capquanli, input_lop_truc, input_trangthai, account_id))
            conn.commit()
            flash(f"Cập nhật tài khoản '{input_username}' thành công!", 'success')
            return redirect(url_for('index', tuan=selected_tuan))

        else:
            sql_select_edit = "SELECT id, Name, username, password, role, lop, tuan, Capquanli, lop_truc, trangthai FROM accounts WHERE id = %s"
            cursor.execute(sql_select_edit, (account_id,))
            account_to_edit = cursor.fetchone()

            if not account_to_edit:
                flash("Không tìm thấy tài khoản để chỉnh sửa.", 'error')
                return redirect(url_for('index', tuan=selected_tuan))

    except mysql.connector.Error as err:
        flash(f"Lỗi cơ sở dữ liệu: {err}", 'error')
        print(f"Error: {err}")
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()
    
    return render_template('insert_account.html', account_to_edit=account_to_edit, accounts=accounts, selected_tuan=selected_tuan)

@app.route('/delete_account/<int:account_id>')
@login_required
def delete_account(account_id):
    if session.get('role') != 'admin':
        flash("Bạn không có quyền xóa tài khoản.", 'error')
        return redirect(url_for('index'))

    conn = None
    selected_tuan = request.args.get('tuan', type=int)

    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        cursor.execute("DELETE FROM accounts WHERE id = %s", (account_id,))
        conn.commit()
        flash(f"Xóa tài khoản ID {account_id} thành công!", 'success')

    except mysql.connector.Error as err:
        flash(f"Lỗi khi xóa tài khoản: {err}", 'error')
        print(f"Error: {err}")
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

    return redirect(url_for('index', tuan=selected_tuan))
@app.route('/delete_all_accounts', methods=['POST'])
def delete_all_accounts():
    # Chỉ cho admin
    if session.get('role') != 'admin':
        flash('Bạn không có quyền xóa tất cả tài khoản.', 'error')
        return redirect(url_for('index'))

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM accounts")
    conn.commit()
    cursor.close()
    conn.close()

    flash('Đã xóa tất cả tài khoản thành công.', 'success')
    return redirect(url_for('index'))

@app.route('/set_all_tuan', methods=['POST'])
@login_required
def set_all_tuan():
    if session.get('role') != 'admin':
        flash("Bạn không có quyền thực hiện thao tác này.", 'error')
        return redirect(url_for('index'))

    new_tuan = request.form.get('new_tuan_value', type=int)
    
    if new_tuan is None or not (1 <= new_tuan <= 40):
        flash("Giá trị tuần không hợp lệ. Vui lòng chọn tuần từ 1 đến 40.", 'error')
        return redirect(url_for('index'))

    conn = None
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        sql_update_tuan = "UPDATE accounts SET tuan = %s"
        cursor.execute(sql_update_tuan, (new_tuan,))
        
        sql_update_trangthai = "UPDATE accounts SET trangthai = 'Chưa tổng kết'"
        cursor.execute(sql_update_trangthai)

        conn.commit()
        flash(f"Đã cập nhật tất cả tài khoản sang Tuần {new_tuan} và đặt trạng thái tổng kết về 'Chưa tổng kết' thành công!", 'success')

    except mysql.connector.Error as err:
        flash(f"Lỗi khi cập nhật tuần hàng loạt hoặc trạng thái: {err}", 'error')
        print(f"Error: {err}")
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()
    
    return redirect(url_for('index', tuan=new_tuan))

@app.route('/update_lop_truc', methods=['POST'])
@login_required
def update_lop_truc_route():
    if session.get('role') != 'admin':
        flash("Bạn không có quyền cập nhật lớp trực.", 'error')
        return redirect(url_for('index'))

    message = update_lop_truc_data()
    flash(message)
    return redirect(url_for('index'))

@app.route('/save_schedule', methods=['POST'])
@login_required
def save_schedule():
    if session.get('role') != 'admin':
        return jsonify({'message': 'Bạn không có quyền thực hiện hành động này.'}), 403

    data = request.get_json()
    try:
        insert_schedule(data)
        return jsonify({'message': 'Đã lưu vào SQL thành công!'}), 200
    except Exception as e:
        return jsonify({'message': f'Lỗi khi lưu lịch: {str(e)}'}), 500

@app.route('/update_schedule', methods=['POST'])
@login_required
def update_schedule_route():
    if session.get('role') != 'admin':
        return jsonify({'message': 'Bạn không có quyền thực hiện hành động này.'}), 403

    data = request.get_json()
    try:
        update_schedule(data)
        return jsonify({'message': 'Đã cập nhật phân công thành công!'}), 200
    except Exception as e:
        return jsonify({'message': f'Lỗi khi cập nhật lịch: {str(e)}'}), 500

@app.route('/clear_all', methods=['POST'])
@login_required
def clear_all_route():
    if session.get('role') != 'admin':
        return jsonify({'message': 'Bạn không có quyền thực hiện hành động này.'}), 403

    try:
        clear_all_schedule()
        return jsonify({'message': 'Đã xóa tất cả phân công trong cơ sở dữ liệu!'}), 200
    except Exception as e:
        return jsonify({'message': f'Lỗi khi xóa lịch: {str(e)}'}), 500

@app.route('/api/get_class_summary_status', methods=['GET'])
@login_required
def get_class_summary_status():
    conn = None
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT MAX(tuan) AS latest_tuan FROM bang_tong_ket")
        result = cursor.fetchone()
        latest_tuan = result['latest_tuan']

        if not latest_tuan:
            return jsonify({'class_statuses': [], 'current_week': None, 'message': 'Không có dữ liệu tổng kết nào để hiển thị trạng thái lớp.'}), 200

        cursor.execute("SELECT DISTINCT lop FROM bang_tong_ket WHERE tuan = %s", (latest_tuan,))
        duty_classes_raw = cursor.fetchall()
        duty_classes = [row['lop'] for row in duty_classes_raw]

        cursor.execute("SELECT username, trangthai FROM accounts WHERE tuan = %s", (latest_tuan,))
        user_statuses = {row['username']: row['trangthai'] for row in cursor.fetchall()}

        cursor.execute("SELECT username, lop FROM info_data")
        user_classes = {row['username']: row['lop'] for row in cursor.fetchall()}

        class_summary_data = []
        for duty_class in duty_classes:
            has_summarized = False
            for username, user_class in user_classes.items():
                if user_class == duty_class and user_statuses.get(username) == 'Đã tổng kết':
                    has_summarized = True
                    break
            
            status_text = 'Đã tổng kết' if has_summarized else 'Chưa tổng kết'
            class_summary_data.append({
                'lop': duty_class,
                'status': status_text,
                'tuan': latest_tuan
            })
        
        class_summary_data.sort(key=lambda x: x['lop'])

        return jsonify({'class_statuses': class_summary_data, 'current_week': latest_tuan})

    except mysql.connector.Error as err:
        print(f"Lỗi MySQL khi lấy trạng thái tổng kết lớp: {err}")
        return jsonify({'error': f'Lỗi khi lấy trạng thái tổng kết lớp: {str(err)}'}), 500
    except Exception as e:
        print(f"Lỗi không xác định khi lấy trạng thái tổng kết lớp: {e}")
        return jsonify({'error': f'Lỗi không xác định: {str(e)}'}), 500
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/add_accounts_from_excel', methods=['POST'])
@login_required
def add_accounts_from_excel():
    if session.get('role') != 'admin':
        flash("Bạn không có quyền thực hiện hành động này.", 'error')
        return redirect(url_for('index'))

    file = request.files.get('file')
    if not file:
        flash("Vui lòng chọn file Excel.", 'error')
        return redirect(url_for('index'))

    filepath = os.path.join('uploads', secure_filename(file.filename))
    os.makedirs('uploads', exist_ok=True)
    file.save(filepath)

    try:
        wb = load_workbook(filepath)
        sheet = wb.active

        conn = get_db_connection()
        cursor = conn.cursor()
        added_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, username, lop, capquanli = row
            if not (name and username and lop and capquanli):
                continue

            # Sinh mật khẩu tự động
            password = generate_specific_password()
            tuan = 1
            trangthai = "Chưa tổng kết"
            lop_truc = ""

            try:
                cursor.execute("""
                    INSERT INTO accounts (Name, username, password, role, lop, tuan, Capquanli, lop_truc, trangthai)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (name, username, password, capquanli, lop, tuan, capquanli, lop_truc, trangthai))
                added_count += 1
            except mysql.connector.IntegrityError:
                # Trùng username thì bỏ qua
                continue

        conn.commit()
        flash(f"Đã thêm {added_count} tài khoản từ file Excel.", 'success')

    except Exception as e:
        flash(f"Lỗi khi xử lý file Excel: {e}", 'error')
    finally:
        cursor.close()
        conn.close()
        os.remove(filepath)

    return redirect(url_for('index'))





@app.route('/user_tong_ket', methods=['POST'])
@login_required
def user_tong_ket():
    if session.get('role') != 'user':
        flash("Bạn không có quyền thực hiện thao tác này.", 'error')
        return redirect(url_for('home'))
    user_username = session.get('username')
    user_lop = session.get('lop')
    user_tuan = session.get('tuan')
    user_lop_truc = session.get('lop_truc')
    conn = None
    cursor = None
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True, buffered=True)
        cursor.execute("SELECT trangthai FROM bang_tong_ket WHERE lop_truc = %s AND tuan = %s", (user_lop_truc, user_tuan))
        existing_summary = cursor.fetchone()
        if existing_summary and existing_summary['trangthai'] == 'Đã tổng kết':
            flash(f"Tuần {user_tuan} đã được tổng kết. Bạn không thể tổng kết lại.", 'warning')
            return redirect(url_for('home'))
        cursor.execute("SELECT SUM(tong_diem) as study FROM study_data WHERE lop_truc = %s AND tuan = %s", (user_lop_truc, user_tuan))
        result_study = cursor.fetchone()
        tong_diem_hoc_tap = result_study['study'] if result_study and result_study['study'] is not None else 0
        cursor.execute("SELECT SUM(tong_diem_vi_pham) as rules FROM rules_data WHERE lop_truc = %s AND tuan = %s", (user_lop_truc, user_tuan))
        result_rules = cursor.fetchone()
        tong_diem_noi_quy = result_rules['rules'] if result_rules and result_rules['rules'] is not None else 0
        tong_diem_chung = tong_diem_hoc_tap + tong_diem_noi_quy
        cursor.execute("""
            INSERT INTO bang_tong_ket 
            (tuan, lop_truc, tong_diem_hoc_tap, tong_diem_noi_quy, tong_diem_chung, trangthai)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE 
                tong_diem_hoc_tap = %s, 
                tong_diem_noi_quy = %s, 
                tong_diem_chung = %s,
                trangthai = %s
        """, (
            user_tuan, user_lop_truc, tong_diem_hoc_tap, tong_diem_noi_quy, tong_diem_chung, 'Đã tổng kết',
            tong_diem_hoc_tap, tong_diem_noi_quy, tong_diem_chung, 'Đã tổng kết'
        ))
        cursor.execute("""
            UPDATE accounts 
            SET trangthai = 'Đã tổng kết'
            WHERE username = %s AND tuan = %s
        """, (user_username, user_tuan))
        conn.commit()
        flash(f"Đã tổng kết thành công cho lớp trực {user_lop_truc} – Tuần {user_tuan}.", 'success')
    except mysql.connector.Error as err:
        flash(f"Lỗi khi tổng kết: {err}", 'error')
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
    return redirect(url_for('home'))

@app.route('/get_accounts_status', methods=['GET'])
@login_required
def get_accounts_status():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Bạn không có quyền xem thông tin này.'}), 403

    conn = None
    accounts_data = []
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT username, lop_truc, trangthai, tuan FROM accounts")
        accounts_data = cursor.fetchall()
    except mysql.connector.Error as err:
        print(f"Error fetching accounts status: {err}")
        return jsonify({'error': f"Lỗi khi lấy dữ liệu tài khoản: {err}"}), 500
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()
    return jsonify({'accounts': accounts_data})

@app.route('/reset_account_status', methods=['POST'])
@login_required
def reset_account_status():
    if session.get('role') != 'admin':
        flash("Bạn không có quyền thực hiện thao tác này.", 'error')
        return redirect(url_for('home'))

    data = request.get_json()
    username_to_reset = data.get('username')
    tuan_to_reset = data.get('tuan')

    if not username_to_reset or not tuan_to_reset:
        return jsonify({'error': 'Thiếu thông tin người dùng hoặc tuần.'}), 400

    conn = None
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute("UPDATE accounts SET trangthai = 'Chưa tổng kết' WHERE username = %s AND tuan = %s", (username_to_reset, tuan_to_reset))
        conn.commit()
        return jsonify({'message': f"Đã đặt lại trạng thái cho {username_to_reset} Tuần {tuan_to_reset} thành 'Chưa tổng kết'."}), 200
    except mysql.connector.Error as err:
        print(f"Error resetting account status: {err}")
        return jsonify({'error': f"Lỗi khi đặt lại trạng thái: {err}"}), 500
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

@app.route("/phancong", methods=["GET", "POST"])
@login_required
def phancong_index():
    if session.get('role') != 'admin':
        flash("Bạn không có quyền truy cập tính năng phân công.", 'error')
        return redirect(url_for('home'))

    result = []
    if request.method == "POST":
        khoi = request.form.get("khoi", "10")
        so_lop = int(request.form.get("so_lop", 20))
        so_tuan = int(request.form.get("so_tuan", 21))
        danh_sach_lop = [f"{khoi}A{i}" for i in range(1, so_lop + 1)]
        print(f"Phân công: khoi={khoi}, so_lop={so_lop}, so_tuan={so_tuan}, danh_sach_lop={danh_sach_lop}")  # Debug

        for i in range(so_lop):
            lop_hien_tai = danh_sach_lop[i]
            phan_cong = []
            j = 1
            while len(phan_cong) < so_tuan:
                index = (i + j) % so_lop
                lop_truc = danh_sach_lop[index]
                if lop_truc != lop_hien_tai:
                    phan_cong.append(lop_truc)
                j += 1
            print(f"Lớp {lop_hien_tai}: {phan_cong}")  # Debug
            save_phancong(khoi, lop_hien_tai, phan_cong)
            result.append({
                "khoi": khoi,
                "lop": lop_hien_tai,
                "phan_cong": phan_cong
            })
        update_lop_truc_data()  # Cập nhật lop_truc sau khi phân công
        flash("Phân công trực lớp và cập nhật tài khoản thành công!", 'success')
    else:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT khoi, lop, tuan, lop_truc FROM phan_cong_truc ORDER BY lop, tuan")
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        tmp = defaultdict(lambda: {"phan_cong": []})
        for khoi, lop, tuan, lop_truc in rows:
            tmp[lop]["khoi"] = khoi
            tmp[lop]["lop"] = lop
            tmp[lop]["phan_cong"].append(lop_truc)

        for lop, data in tmp.items():
            result.append(data)

    return render_template("testphancong.html", result=result)

@app.route('/clear_phancong_data', methods=['POST'])
@login_required
def clear_phancong_data():
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Bạn không có quyền thực hiện hành động này.'}), 403

    conn = None
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute("TRUNCATE TABLE phan_cong_truc")
        conn.commit()
        flash("Đã xóa toàn bộ dữ liệu phân công trực thành công.", 'success')
        return jsonify({'status': 'success', 'message': 'Đã xóa toàn bộ dữ liệu phân công trực thành công.'}), 200
    except mysql.connector.Error as err:
        flash(f"Lỗi khi xóa dữ liệu phân công trực: {err}", 'error')
        print(f"Error clearing phancong_truc: {err}")
        return jsonify({'status': 'error', 'message': f'Lỗi khi xóa dữ liệu: {err}'}), 500
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/edit_rule', methods=['POST'])
@login_required
def edit_rule():
    rule_id = request.form['id']
    new_content = request.form['content']
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE rules_data SET content=%s WHERE id=%s", (new_content, rule_id))
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify(success=True)

@app.route('/delete_rule', methods=['POST'])
@login_required
def delete_rule():
    if session.get('role') not in ['admin', 'user']:
        flash("Bạn không có quyền xóa dữ liệu nội quy.", 'error')
        return jsonify({'status': 'error', 'message': 'Bạn không có quyền xóa dữ liệu nội quy.'}), 403

    rule_id = request.form['id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT tuan FROM rules_data WHERE id = %s", (rule_id,))
    rule_info = cursor.fetchone()
    if not rule_info:
        cursor.close()
        conn.close()
        return jsonify({'status': 'error', 'message': 'Dữ liệu nội quy không tồn tại.'}), 404

    cursor.execute("SELECT trangthai FROM accounts WHERE username = %s AND tuan = %s", 
                   (session.get('username'), rule_info['tuan']))
    account_status = cursor.fetchone()
    if account_status and account_status['trangthai'] == 'Đã tổng kết':
        flash("Không thể xóa dữ liệu nội quy vì tuần này đã được tổng kết.", 'error')
        cursor.close()
        conn.close()
        return jsonify({'status': 'error', 'message': 'Không thể xóa dữ liệu nội quy vì tuần này đã được tổng kết.'}), 403

    cursor.execute("DELETE FROM rules_data WHERE id=%s", (rule_id,))
    conn.commit()
    flash("Đã xóa dữ liệu nội quy thành công.", 'success')
    cursor.close()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Đã xóa dữ liệu nội quy thành công.'})

@app.route('/update_study_data/<int:data_id>', methods=['GET', 'POST'])
@login_required
def update_study_data(data_id):
    if session.get('role') != 'user' and session.get('role') != 'giamthi' and session.get('role') != 'admin':
        flash("Bạn không có quyền chỉnh sửa dữ liệu học tập của trang người dùng.", 'error')
        return redirect(url_for('user'))
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT tuan, lop_truc FROM study_data WHERE id = %s", (data_id,))
    study_data = cursor.fetchone()
    if not study_data:
        flash("Không tìm thấy dữ liệu học tập.", 'error')
        cursor.close()
        conn.close()
        return redirect(url_for('user'))
    cursor.execute("SELECT trangthai FROM bang_tong_ket WHERE lop_truc = %s AND tuan = %s", (study_data['lop_truc'], study_data['tuan']))
    account_status = cursor.fetchone()
    if account_status and account_status['trangthai'] == 'Đã tổng kết':
        flash("Không thể chỉnh sửa dữ liệu học tập vì tuần này đã được tổng kết.", 'warning')
        cursor.close()
        conn.close()
        return redirect(url_for('user'))
    if request.method == 'POST':
        gio_a = int(request.form.get('gio_a') or 0)
        gio_b = int(request.form.get('gio_b') or 0)
        gio_c = int(request.form.get('gio_c') or 0)
        gio_d = int(request.form.get('gio_d') or 0)
        dat_kieu_mau = request.form['dat_kieu_mau']
        tong_diem = gio_a * 5 + gio_b * -5 + gio_c * -15 + gio_d * -25
        tong_diem += 5 if dat_kieu_mau == "Yes" else -10
        try:
            cursor.execute("""
                UPDATE study_data 
                SET gio_a = %s, gio_b = %s, gio_c = %s, gio_d = %s, dat_kieu_mau = %s, tong_diem = %s
                WHERE id = %s
            """, (gio_a, gio_b, gio_c, gio_d, dat_kieu_mau, tong_diem, data_id))
            conn.commit()
            flash("Đã cập nhật dữ liệu học tập thành công.", 'success')
        except mysql.connector.Error as err:
            flash(f"Lỗi khi cập nhật dữ liệu học tập: {err}", 'error')
        finally:
            cursor.close()
            conn.close()
        return redirect(url_for('user'))
    cursor.execute("SELECT * FROM study_data WHERE id = %s", (data_id,))
    data = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('update_study_data.html', data=data)



@app.route('/update_rules_data/<int:data_id>', methods=['GET', 'POST'])
@login_required
def update_rules_data(data_id):
    if session.get('role') not in ['user', 'giamthi']:
    
        flash("Bạn không có quyền chỉnh sửa dữ liệu nội quy.", 'error')
        return redirect(url_for('noi_quy'))
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Kiểm tra dữ liệu nội quy
    cursor.execute("SELECT * FROM rules_data WHERE id = %s", (data_id,))
    rules_data = cursor.fetchone()
    if not rules_data:
        flash("Không tìm thấy dữ liệu nội quy.", 'error')
        cursor.close()
        conn.close()
        return redirect(url_for('noi_quy'))
    
    # Kiểm tra quyền chỉnh sửa
    user_lop_truc = session.get('lop_truc')
    user_tuan = session.get('tuan')
    if rules_data['lop_truc'] != user_lop_truc or rules_data['tuan'] != user_tuan:
        flash("Bạn chỉ có thể chỉnh sửa dữ liệu của lớp trực và tuần được gán.", 'error')
        cursor.close()
        conn.close()
        return redirect(url_for('noi_quy'))
    
    # Kiểm tra trạng thái tổng kết
    cursor.execute("SELECT trangthai FROM bang_tong_ket WHERE lop_truc = %s AND tuan = %s", (rules_data['lop_truc'], rules_data['tuan']))
    account_status = cursor.fetchone()
    if account_status and account_status['trangthai'] == 'Đã tổng kết':
        flash("Không thể chỉnh sửa dữ liệu nội quy vì tuần này đã được tổng kết.", 'warning')
        cursor.close()
        conn.close()
        return redirect(url_for('noi_quy'))
    
    if request.method == 'POST':
        noi_dung_vi_pham = request.form.get('noi_dung_vi_pham')
        diem_tru = int(request.form.get('diem_tru') or 0)
        so_luot_vi_pham = int(request.form.get('so_luot_vi_pham') or 0)
        ten_hoc_sinh_vi_pham = request.form.get('ten_hoc_sinh_vi_pham') or ''
        tong_diem_vi_pham = diem_tru * so_luot_vi_pham
        try:
            cursor.execute("""
                UPDATE rules_data 
                SET noi_dung_vi_pham = %s, diem_tru = %s, so_luot_vi_pham = %s, ten_hoc_sinh_vi_pham = %s, tong_diem_vi_pham = %s
                WHERE id = %s
            """, (noi_dung_vi_pham, diem_tru, so_luot_vi_pham, ten_hoc_sinh_vi_pham, tong_diem_vi_pham, data_id))
            conn.commit()
            flash("Đã cập nhật dữ liệu nội quy thành công.", 'success')
        except mysql.connector.Error as err:
            flash(f"Lỗi khi cập nhật dữ liệu nội quy: {err}", 'error')
        finally:
            cursor.close()
            conn.close()
        return redirect(url_for('noi_quy'))
    
    cursor.close()
    conn.close()
    return render_template('update_rules_data.html', data=rules_data)


@app.route('/home_public', methods=['GET'])
def home_public():
    print("Accessing home_public")  # Debug log
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM study_data ORDER BY tuan DESC, lop_truc ASC")
    study_data = cursor.fetchall()

    cursor.execute("SELECT * FROM rules_data ORDER BY tuan DESC, lop_truc ASC")
    rules_data = cursor.fetchall()

    cursor.execute("SELECT DISTINCT tuan FROM bang_tong_ket UNION SELECT DISTINCT tuan FROM study_data UNION SELECT DISTINCT tuan FROM rules_data ORDER BY tuan ASC")
    available_export_weeks = [row['tuan'] for row in cursor.fetchall()]
    
    cursor.execute("SELECT DISTINCT lop_truc FROM bang_tong_ket UNION SELECT DISTINCT lop_truc FROM study_data UNION SELECT DISTINCT lop_truc FROM rules_data ORDER BY lop_truc ASC")
    available_export_classes = [row['lop_truc'] for row in cursor.fetchall()]
    
    cursor.close()
    conn.close()

    return render_template('home.html', study_data=study_data, rules_data=rules_data, 
                           lop='N/A', tuan='N/A', lop_truc='N/A',
                           available_export_weeks=available_export_weeks, 
                           available_export_classes=available_export_classes, 
                           is_public=True)

@app.route('/update_rules_data_admin/<int:data_id>', methods=['GET', 'POST'])
@login_required
def update_rules_data_admin(data_id):
    if session.get('role') not in ['admin', 'giamthi']:
    
        flash("Bạn không có quyền chỉnh sửa dữ liệu nội quy.", 'error')
        return redirect(url_for('noi_quy'))
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Kiểm tra trạng thái tổng kết
    cursor.execute("SELECT tuan, lop_truc FROM rules_data WHERE id = %s", (data_id,))
    rules_data = cursor.fetchone()
    cursor.fetchall()  # Đọc hết kết quả
    if not rules_data:
        flash("Không tìm thấy dữ liệu nội quy.", 'error')
        cursor.close()
        conn.close()
        return redirect(url_for('noi_quy'))
    
    cursor.execute("SELECT trangthai FROM bang_tong_ket WHERE lop_truc = %s AND tuan = %s", (rules_data['lop_truc'], rules_data['tuan']))
    account_status = cursor.fetchone()
    cursor.fetchall()  # Đọc hết kết quả
    if account_status and account_status['trangthai'] == 'Đã tổng kết':
        flash("Không thể chỉnh sửa dữ liệu nội quy vì tuần này đã được tổng kết.", 'warning')
        cursor.close()
        conn.close()
        return redirect(url_for('noi_quy'))
    
    if request.method == 'POST':
        noi_dung_vi_pham = request.form.get('noi_dung_vi_pham')
        diem_tru = int(request.form.get('diem_tru') or 0)
        so_luot_vi_pham = int(request.form.get('so_luot_vi_pham') or 0)
        ten_hoc_sinh_vi_pham = request.form.get('ten_hoc_sinh_vi_pham')
        tong_diem_vi_pham = diem_tru * so_luot_vi_pham
        try:
            cursor.execute("""
                UPDATE rules_data 
                SET noi_dung_vi_pham = %s, diem_tru = %s, so_luot_vi_pham = %s, ten_hoc_sinh_vi_pham = %s, tong_diem_vi_pham = %s
                WHERE id = %s
            """, (noi_dung_vi_pham, diem_tru, so_luot_vi_pham, ten_hoc_sinh_vi_pham, tong_diem_vi_pham, data_id))
            conn.commit()
            flash("Đã cập nhật dữ liệu nội quy thành công.", 'success')
        except mysql.connector.Error as err:
            flash(f"Lỗi khi cập nhật dữ liệu nội quy: {err}", 'error')
        finally:
            cursor.close()
            conn.close()
        return redirect(url_for('noi_quy'))
    
    cursor.execute("SELECT * FROM rules_data WHERE id = %s", (data_id,))
    data = cursor.fetchone()
    cursor.fetchall()  # Đọc hết kết quả
    cursor.close()
    conn.close()
    return render_template('update_rules_data_admin.html', data=data)


@app.route('/toggle_tong_ket/<lop>/<tuan>', methods=['POST'])
@login_required
def toggle_tong_ket(lop, tuan):
    if session.get('role') != 'admin':
        abort(403)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Lấy trạng thái hiện tại
    cursor.execute("SELECT trangthai FROM bang_tong_ket WHERE lop = %s AND tuan = %s", (lop, tuan))
    row = cursor.fetchone()

    if not row:
        flash(f"Không tìm thấy bản ghi tổng kết của {lop} – Tuần {tuan}", 'error')
    else:
        new_status = 'Chưa tổng kết' if row['trangthai'] == 'Đã tổng kết' else 'Đã tổng kết'

        # Cập nhật bang_tong_ket
        cursor.execute("UPDATE bang_tong_ket SET trangthai = %s WHERE lop = %s AND tuan = %s", (new_status, lop, tuan))

        # Cập nhật luôn accounts (nếu bạn vẫn dùng accounts.trangthai để hiển thị nhanh)
        cursor.execute("UPDATE accounts SET trangthai = %s WHERE lop = %s AND tuan = %s", (new_status, lop, tuan))

        conn.commit()
        flash(f"Đã cập nhật trạng thái tổng kết của {lop} – Tuần {tuan} thành: {new_status}", 'success')

    cursor.close()
    conn.close()
    return redirect(url_for('admin'))

@app.route('/tong_ket_tat_ca', methods=['GET'])
@login_required
def tong_ket_tat_ca():
    if session.get('role') != 'admin':
        flash("Bạn không có quyền truy cập trang tổng kết tất cả.", 'error')
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Lấy danh sách tuần, khối
    cursor.execute("SELECT DISTINCT tuan FROM bang_tong_ket ORDER BY tuan ASC")
    available_weeks = [row['tuan'] for row in cursor.fetchall()]
    cursor.execute("SELECT DISTINCT khoi FROM bang_tong_ket WHERE khoi IS NOT NULL ORDER BY khoi ASC")
    available_khoi = [row['khoi'] for row in cursor.fetchall()]
    
    selected_tuan = request.args.get('tuan', type=str)
    selected_khoi = request.args.get('khoi', type=str)
    view_all = request.args.get('view_all', type=str) == 'true'
    
    # Lấy dữ liệu tổng kết với các cột chi tiết
    query = """
        SELECT bt.tuan, bt.khoi, bt.lop_truc, 
               COALESCE(SUM(sd.gio_a), 0) as gio_a, 
               COALESCE(SUM(sd.gio_b), 0) as gio_b, 
               COALESCE(SUM(sd.gio_c), 0) as gio_c, 
               COALESCE(SUM(sd.gio_d), 0) as gio_d,
               COALESCE(SUM(rd.so_luot_vi_pham), 0) as so_luot_vi_pham,
               GROUP_CONCAT(DISTINCT rd.noi_dung_vi_pham SEPARATOR ', ') as noi_dung_vi_pham,
               GROUP_CONCAT(DISTINCT rd.ten_hoc_sinh_vi_pham SEPARATOR ', ') as hoc_sinh_vi_pham,
               bt.tong_diem_hoc_tap, bt.tong_diem_noi_quy, bt.tong_diem_chung
        FROM bang_tong_ket bt
        LEFT JOIN study_data sd ON bt.tuan = sd.tuan AND bt.lop_truc = sd.lop_truc
        LEFT JOIN rules_data rd ON bt.tuan = rd.tuan AND bt.lop_truc = rd.lop_truc
        WHERE 1=1
    """
    params = []
    if selected_tuan and not view_all:
        query += " AND bt.tuan = %s"
        params.append(selected_tuan)
    if selected_khoi:
        query += " AND bt.khoi = %s"
        params.append(selected_khoi)
    query += " GROUP BY bt.tuan, bt.khoi, bt.lop_truc, bt.tong_diem_hoc_tap, bt.tong_diem_noi_quy, bt.tong_diem_chung"
    query += " ORDER BY bt.tuan ASC, bt.khoi ASC, bt.tong_diem_chung DESC"
    
    cursor.execute(query, params)
    data = cursor.fetchall()
    cursor.fetchall()  # Đọc hết kết quả
    
    if not data:
        flash("Không có dữ liệu tổng kết để hiển thị.", 'warning')
        cursor.close()
        conn.close()
        return render_template('tong_ket_tat_ca.html', data=[], available_weeks=available_weeks, available_khoi=available_khoi, selected_tuan=selected_tuan, selected_khoi=selected_khoi)
    
    # Tính toán xếp hạng
    ranked_data = []
    grouped_data = defaultdict(list)
    for item in data:
        grouped_data[(item['tuan'], item.get('khoi', 'Unknown'))].append(item)
    for tuan_khoi in sorted(grouped_data.keys()):
        current_week_block_data = sorted(grouped_data[tuan_khoi], key=lambda x: x['tong_diem_chung'], reverse=True)
        current_rank = 1
        same_rank_count = 1
        prev_diem = None
        for i, item in enumerate(current_week_block_data):
            if i == 0:
                item['xep_hang'] = current_rank
            else:
                if item['tong_diem_chung'] < prev_diem:
                    current_rank += same_rank_count
                    same_rank_count = 1
                else:
                    same_rank_count += 1
                item['xep_hang'] = current_rank
            prev_diem = item['tong_diem_chung']
            ranked_data.append(item)
    
    cursor.close()
    conn.close()
    return render_template('tong_ket_tat_ca.html', data=ranked_data, available_weeks=available_weeks, available_khoi=available_khoi, selected_tuan=selected_tuan, selected_khoi=selected_khoi)

@app.route('/export_tong_ket_tat_ca', methods=['GET'])
@login_required
def export_tong_ket_tat_ca():
    if session.get('role') != 'admin':
        flash("Bạn không có quyền xuất dữ liệu tổng kết tất cả.", 'error')
        return redirect(url_for('home'))
    
    selected_tuan = request.args.get('tuan', type=str)
    selected_khoi = request.args.get('khoi', type=str)
    view_all = request.args.get('view_all', type=str) == 'true'
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    query = """
        SELECT bt.tuan, bt.khoi, bt.lop_truc, 
               COALESCE(SUM(sd.gio_a), 0) as gio_a, 
               COALESCE(SUM(sd.gio_b), 0) as gio_b, 
               COALESCE(SUM(sd.gio_c), 0) as gio_c, 
               COALESCE(SUM(sd.gio_d), 0) as gio_d,
               COALESCE(SUM(rd.so_luot_vi_pham), 0) as so_luot_vi_pham,
               GROUP_CONCAT(DISTINCT rd.noi_dung_vi_pham SEPARATOR ', ') as noi_dung_vi_pham,
               GROUP_CONCAT(DISTINCT rd.ten_hoc_sinh_vi_pham SEPARATOR ', ') as hoc_sinh_vi_pham,
               bt.tong_diem_hoc_tap, bt.tong_diem_noi_quy, bt.tong_diem_chung
        FROM bang_tong_ket bt
        LEFT JOIN study_data sd ON bt.tuan = sd.tuan AND bt.lop_truc = sd.lop_truc
        LEFT JOIN rules_data rd ON bt.tuan = rd.tuan AND bt.lop_truc = rd.lop_truc
        WHERE 1=1
    """
    params = []
    if selected_tuan and not view_all:
        query += " AND bt.tuan = %s"
        params.append(selected_tuan)
    if selected_khoi:
        query += " AND bt.khoi = %s"
        params.append(selected_khoi)
    query += " GROUP BY bt.tuan, bt.khoi, bt.lop_truc, bt.tong_diem_hoc_tap, bt.tong_diem_noi_quy, bt.tong_diem_chung"
    query += " ORDER BY bt.tuan ASC, bt.khoi ASC, bt.tong_diem_chung DESC"
    
    cursor.execute(query, params)
    data = cursor.fetchall()
    cursor.fetchall()  # Đọc hết kết quả
    cursor.close()
    conn.close()
    
    if not data:
        flash("Không có dữ liệu tổng kết để xuất.", 'warning')
        return redirect(url_for('tong_ket_tat_ca'))
    
    # Tính toán xếp hạng
    ranked_data = []
    grouped_data = defaultdict(list)
    for item in data:
        grouped_data[(item['tuan'], item.get('khoi', 'Unknown'))].append(item)
    for tuan_khoi in sorted(grouped_data.keys()):
        current_week_block_data = sorted(grouped_data[tuan_khoi], key=lambda x: x['tong_diem_chung'], reverse=True)
        current_rank = 1
        same_rank_count = 1
        prev_diem = None
        for i, item in enumerate(current_week_block_data):
            if i == 0:
                item['xep_hang'] = current_rank
            else:
                if item['tong_diem_chung'] < prev_diem:
                    current_rank += same_rank_count
                    same_rank_count = 1
                else:
                    same_rank_count += 1
                item['xep_hang'] = current_rank
            prev_diem = item['tong_diem_chung']
            ranked_data.append(item)
    
    # Tạo DataFrame với tên cột tiếng Việt
    df = pd.DataFrame(ranked_data, columns=[
        'tuan', 'khoi', 'lop_truc', 'gio_a', 'gio_b', 'gio_c', 'gio_d', 
        'so_luot_vi_pham', 'noi_dung_vi_pham', 'hoc_sinh_vi_pham', 
        'tong_diem_hoc_tap', 'tong_diem_noi_quy', 'tong_diem_chung', 'xep_hang'
    ])
    df = df.rename(columns={
        'tuan': 'Tuần',
        'khoi': 'Khối',
        'lop_truc': 'Lớp',
        'gio_a': 'Giờ A',
        'gio_b': 'Giờ B',
        'gio_c': 'Giờ C',
        'gio_d': 'Giờ D',
        'so_luot_vi_pham': 'Số Lượt Vi Phạm',
        'noi_dung_vi_pham': 'Nội Dung Vi Phạm',
        'hoc_sinh_vi_pham': 'Học Sinh Vi Phạm',
        'tong_diem_hoc_tap': 'Tổng Điểm Học Tập',
        'tong_diem_noi_quy': 'Tổng Điểm Nội Quy',
        'tong_diem_chung': 'Tổng Điểm Chung',
        'xep_hang': 'Xếp Hạng'
    })
    
    # Xuất Excel với định dạng thẩm mỹ
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Tổng Kết Tất Cả', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Tổng Kết Tất Cả']
        
        # Định dạng tiêu đề
        header_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#1e40af',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'left',
            'valign': 'vcenter'
        })
        
        # Áp dụng định dạng tiêu đề
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
        
        # Áp dụng định dạng cho các ô dữ liệu
        for row_num in range(1, len(df) + 1):
            for col_num in range(len(df.columns)):
                value = df.iloc[row_num - 1, col_num]
                worksheet.write(row_num, col_num, value if value is not None else 'Không có', cell_format)
        
        # Tự động điều chỉnh độ rộng cột
        for col_num, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
            worksheet.set_column(col_num, col_num, max_len)
    
    output.seek(0)
    filename = f"tong_ket_tat_ca_{selected_tuan or 'tat_ca'}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/toggle_tong_ket_status', methods=['POST'])
@login_required
def toggle_tong_ket_status():
    if session.get('role') != 'admin':
        abort(403)

    lop = request.form.get('lop')
    tuan = request.form.get('tuan')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)  # ✅ buffer nếu cần nhiều truy vấn

    # ✅ BẮT BUỘC: đọc kết quả trước khi thực hiện truy vấn khác
    cursor.execute("SELECT trangthai FROM bang_tong_ket WHERE lop = %s AND tuan = %s", (lop, tuan))
    result = cursor.fetchone()

    if result:
        # ✅ Đọc xong mới được xóa
        cursor.execute("DELETE FROM bang_tong_ket WHERE lop = %s AND tuan = %s", (lop, tuan))
        cursor.execute("UPDATE accounts SET trangthai = 'Chưa tổng kết' WHERE lop = %s AND tuan = %s", (lop, tuan))
        conn.commit()
        flash(f"Đã tổng kết lại lớp {lop} – Tuần {tuan}. Dữ liệu đã bị xoá.", 'success')
    else:
        flash(f"Không có dữ liệu để tổng kết lại.", 'info')

    cursor.close()
    conn.close()

    return redirect(url_for('tong_ket', tuan=tuan))

if __name__ == '__main__':
    create_table()  # Tạo bảng phan_cong_truc nếu chưa tồn tại
    app.run(host='0.0.0.0', port=5000, debug=True)